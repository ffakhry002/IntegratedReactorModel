"""
Excel report generation for test results
UPDATED: Support for different flux modes (total, energy, bin)
"""

import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelReporter:
    """Class to handle Excel report generation"""

    def create_report(self, results, output_filename):
        """Create comprehensive Excel report with test results"""
        # Convert results to DataFrame
        df = pd.DataFrame(results)

                # Determine which types of results we have
        has_keff = any('keff_real' in r for r in results if r)
        # Check for any flux columns - could be I_1_real, I_1_fast_real, I_1_thermal_real, etc.
        has_flux = any(any('I_1' in key and '_real' in key for key in r.keys()) for r in results if r)

        # Detect flux mode
        flux_mode = 'total'  # default
        if any('flux_mode' in r for r in results if r):
            # Get the flux mode from first flux result
            flux_results = [r for r in results if r and r.get('model_type') == 'flux']
            if flux_results:
                flux_mode = flux_results[0].get('flux_mode', 'total')

        if results:
            sample_keys = list(results[0].keys())
            fast_keys = [k for k in sample_keys if 'fast' in k]

        # Create Excel writer
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # Create only the main results sheet
            self._create_results_sheet(writer, df, has_keff, has_flux, flux_mode)

        print(f"\nExcel report saved to: {output_filename}")
        print(f"Sheet created:")
        print(f"  - Test Results: Detailed results for all configurations")
        if flux_mode != 'total':
            print(f"    (Flux mode: {flux_mode} - showing energy-discretized results)")

        return output_filename

    def _create_results_sheet(self, writer, df, has_keff, has_flux, flux_mode='total'):
        """Create the main results sheet with merged keff and flux rows"""

        # Define the key columns for matching rows (excluding in_training)
        key_columns = ['config_id', 'description', 'model_class', 'encoding', 'optimization_method']

        # Add flux_mode column if we have energy/bin modes
        if flux_mode != 'total' and 'flux_mode' in df.columns:
            key_columns.append('flux_mode')

        # If we have both keff and flux, merge them
        if has_keff and has_flux:
            # Separate keff and flux rows
            keff_df = df[df['model_type'] == 'keff'].copy()
            flux_df = df[df['model_type'] == 'flux'].copy()

            # Drop model_type from both as it's no longer needed after merge
            keff_df = keff_df.drop('model_type', axis=1, errors='ignore')
            flux_df = flux_df.drop('model_type', axis=1, errors='ignore')

            # Identify keff specific columns
            keff_specific_cols = ['keff_real', 'keff_predicted', 'keff_rel_error']
            if 'mape_keff' in keff_df.columns:
                keff_specific_cols.append('mape_keff')

            # Identify flux specific columns based on flux mode
            flux_specific_cols = []

            if flux_mode == 'total':
                # Original total flux columns
                for i in range(1, 5):
                    flux_specific_cols.extend([f'I_{i}_real', f'I_{i}_predicted', f'I_{i}_rel_error'])
                flux_specific_cols.extend(['avg_flux_real', 'avg_flux_predicted', 'avg_flux_rel_error'])

            elif flux_mode in ['thermal_only', 'epithermal_only', 'fast_only']:
                # Single energy group modes
                energy_group = flux_mode.replace('_only', '')
                for i in range(1, 5):
                    flux_specific_cols.extend([
                        f'I_{i}_{energy_group}_real',
                        f'I_{i}_{energy_group}_predicted',
                        f'I_{i}_{energy_group}_rel_error'
                    ])
                # Add average columns for single energy modes
                flux_specific_cols.extend([
                    f'avg_{energy_group}_flux_real',
                    f'avg_{energy_group}_flux_predicted',
                    f'avg_{energy_group}_flux_rel_error',
                    f'mape_{energy_group}_flux'
                ])

            else:  # energy or bin mode (multi-energy)
                # Energy-discretized columns
                energy_groups = ['thermal', 'epithermal', 'fast']
                for i in range(1, 5):
                    for energy in energy_groups:
                        flux_specific_cols.extend([
                            f'I_{i}_{energy}_real',
                            f'I_{i}_{energy}_predicted',
                            f'I_{i}_{energy}_rel_error'
                        ])
                    # Add total columns
                    flux_specific_cols.extend([
                        f'I_{i}_total_real',
                        f'I_{i}_total_predicted',
                        f'I_{i}_total_rel_error'
                    ])

            if 'mape_flux' in flux_df.columns:
                flux_specific_cols.append('mape_flux')

            # Keep only relevant columns for each dataframe
            # Include in_training in the columns to keep
            keff_cols_to_keep = ['in_training'] + key_columns + [col for col in keff_specific_cols if col in keff_df.columns]
            flux_cols_to_keep = ['in_training'] + key_columns + [col for col in flux_specific_cols if col in flux_df.columns]

            keff_df = keff_df[keff_cols_to_keep]
            flux_df = flux_df[flux_cols_to_keep]

            # Merge on key columns AND in_training to ensure they match
            merge_columns = ['in_training'] + key_columns
            merged_df = pd.merge(keff_df, flux_df, on=merge_columns, how='outer')

            # Define column order with in_training first
            base_columns = ['in_training'] + key_columns
            keff_columns = [col for col in keff_specific_cols if col in merged_df.columns]
            flux_columns = [col for col in flux_specific_cols if col in merged_df.columns]

            column_order = base_columns + keff_columns + flux_columns
            column_order = [col for col in column_order if col in merged_df.columns]
            df_ordered = merged_df[column_order]

        else:
            # If we only have one type, use the original logic
            # Put in_training first, then other base columns
            base_columns = ['in_training', 'config_id', 'description', 'model_class', 'model_type', 'encoding', 'optimization_method']

            if flux_mode != 'total' and 'flux_mode' in df.columns:
                base_columns.append('flux_mode')

            keff_columns = []
            if has_keff:
                keff_columns = ['keff_real', 'keff_predicted', 'keff_rel_error']
                if 'mape_keff' in df.columns:
                    keff_columns.append('mape_keff')

            flux_columns = []
            if has_flux:
                if flux_mode == 'total':
                    for i in range(1, 5):
                        flux_columns.extend([f'I_{i}_real', f'I_{i}_predicted', f'I_{i}_rel_error'])
                    flux_columns.extend(['avg_flux_real', 'avg_flux_predicted', 'avg_flux_rel_error'])
                elif flux_mode in ['thermal_only', 'epithermal_only', 'fast_only']:
                    # Single energy group modes
                    energy_group = flux_mode.replace('_only', '')
                    for i in range(1, 5):
                        flux_columns.extend([
                            f'I_{i}_{energy_group}_real',
                            f'I_{i}_{energy_group}_predicted',
                            f'I_{i}_{energy_group}_rel_error'
                        ])
                    # Add average columns for single energy modes
                    flux_columns.extend([
                        f'avg_{energy_group}_flux_real',
                        f'avg_{energy_group}_flux_predicted',
                        f'avg_{energy_group}_flux_rel_error',
                        f'mape_{energy_group}_flux'
                    ])
                else:  # energy or bin mode (multi-energy)
                    energy_groups = ['thermal', 'epithermal', 'fast']
                    for i in range(1, 5):
                        for energy in energy_groups:
                            flux_columns.extend([
                                f'I_{i}_{energy}_real',
                                f'I_{i}_{energy}_predicted',
                                f'I_{i}_{energy}_rel_error'
                            ])
                        flux_columns.extend([
                            f'I_{i}_total_real',
                            f'I_{i}_total_predicted',
                            f'I_{i}_total_rel_error'
                        ])

                if 'mape_flux' in df.columns:
                    flux_columns.append('mape_flux')

            column_order = base_columns + keff_columns + flux_columns
            column_order = [col for col in column_order if col in df.columns]
            df_ordered = df[column_order]

        # Write to Excel
        df_ordered.to_excel(writer, sheet_name='Test Results', index=False)

        # Format the sheet
        ws = writer.sheets['Test Results']
        self._format_worksheet(ws, df_ordered, has_keff, has_flux, flux_mode)

    def _create_summary_sheet(self, writer, df, has_keff, has_flux, flux_mode='total'):
        """Create summary statistics sheet"""
        summary_data = []

        # Group by model characteristics
        grouping_cols = ['model_class', 'model_type', 'encoding', 'optimization_method']
        if flux_mode != 'total' and 'flux_mode' in df.columns:
            grouping_cols.append('flux_mode')

        for name, group in df.groupby(grouping_cols):
            if flux_mode != 'total' and 'flux_mode' in df.columns:
                model_class, model_type, encoding, opt_method, flux_mode_group = name
            else:
                model_class, model_type, encoding, opt_method = name
                flux_mode_group = 'total'

            summary_row = {
                'Model': model_class,
                'Type': model_type,
                'Encoding': encoding,
                'Optimization': opt_method,
                'Configurations Tested': len(group)
            }

            if flux_mode != 'total':
                summary_row['Flux Mode'] = flux_mode_group

            if has_keff and model_type == 'keff' and 'keff_rel_error' in group:
                errors = group['keff_rel_error'].dropna()
                if len(errors) > 0:
                    summary_row['Avg K-eff Error (%)'] = errors.mean()
                    summary_row['Max K-eff Error (%)'] = errors.max()
                    summary_row['Min K-eff Error (%)'] = errors.min()
                    summary_row['Std K-eff Error (%)'] = errors.std()

            if has_flux and model_type == 'flux':
                if flux_mode == 'total' and 'avg_flux_rel_error' in group:
                    errors = group['avg_flux_rel_error'].dropna()
                    if len(errors) > 0:
                        summary_row['Avg Flux Error (%)'] = errors.mean()
                        summary_row['Max Flux Error (%)'] = errors.max()
                        summary_row['Min Flux Error (%)'] = errors.min()
                        summary_row['Std Flux Error (%)'] = errors.std()
                elif flux_mode in ['thermal_only', 'epithermal_only', 'fast_only']:
                    # For single energy modes, use MAPE
                    if 'mape_flux' in group:
                        mape_values = group['mape_flux'].replace('N/A', np.nan).dropna()
                        if len(mape_values) > 0:
                            # Use absolute values for summary statistics
                            abs_mape_values = abs(mape_values)
                            energy_name = flux_mode.replace('_only', '').title()
                            summary_row[f'Avg {energy_name} MAPE (%)'] = abs_mape_values.mean()
                            summary_row[f'Max {energy_name} MAPE (%)'] = abs_mape_values.max()
                            summary_row[f'Min {energy_name} MAPE (%)'] = abs_mape_values.min()
                            summary_row[f'Std {energy_name} MAPE (%)'] = abs_mape_values.std()
                elif flux_mode in ['energy', 'bin'] and 'mape_flux' in group:
                    # For energy/bin modes, use MAPE
                    mape_values = group['mape_flux'].replace('N/A', np.nan).dropna()
                    if len(mape_values) > 0:
                        # Use absolute values for summary statistics
                        abs_mape_values = abs(mape_values)
                        summary_row['Avg MAPE (%)'] = abs_mape_values.mean()
                        summary_row['Max MAPE (%)'] = abs_mape_values.max()
                        summary_row['Min MAPE (%)'] = abs_mape_values.min()
                        summary_row['Std MAPE (%)'] = abs_mape_values.std()

            summary_data.append(summary_row)

        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Format summary sheet
            ws_summary = writer.sheets['Summary']
            self._format_summary_sheet(ws_summary, summary_df)

    def _create_best_models_sheet(self, writer, df, has_keff, has_flux, flux_mode='total'):
        """Create sheet showing best models per encoding"""
        best_models_data = []

        for encoding in df['encoding'].unique():
            encoding_subset = df[df['encoding'] == encoding]

            # Best k-eff model for this encoding
            if has_keff and len(encoding_subset[encoding_subset['model_type'] == 'keff']) > 0:
                keff_subset = encoding_subset[encoding_subset['model_type'] == 'keff']
                if 'mape_keff' in keff_subset:
                    # Group by model and optimization, then find minimum using absolute values
                    abs_keff_values = abs(keff_subset['mape_keff'])
                    keff_subset_abs = keff_subset.copy()
                    keff_subset_abs['mape_keff'] = abs_keff_values
                    grouped = keff_subset_abs.groupby(['model_class', 'optimization_method'])['mape_keff'].mean()
                    if len(grouped) > 0:
                        best_idx = grouped.idxmin()
                        best_models_data.append({
                            'Encoding': encoding,
                            'Target': 'k-eff',
                            'Best Model': best_idx[0],
                            'Optimization': best_idx[1],
                            'Avg Error (%)': grouped.min()
                        })

            # Best flux model for this encoding
            if has_flux and len(encoding_subset[encoding_subset['model_type'] == 'flux']) > 0:
                flux_subset = encoding_subset[encoding_subset['model_type'] == 'flux']

                # Group by flux mode if applicable
                if flux_mode != 'total' and 'flux_mode' in flux_subset.columns:
                    for mode in flux_subset['flux_mode'].unique():
                        mode_subset = flux_subset[flux_subset['flux_mode'] == mode]
                        if 'mape_flux' in mode_subset:
                            # Replace 'N/A' with NaN for proper calculation
                            mode_subset_clean = mode_subset.copy()
                            mode_subset_clean['mape_flux'] = mode_subset_clean['mape_flux'].replace('N/A', np.nan)

                            # Use absolute values for best model selection
                            mode_subset_clean['mape_flux'] = abs(mode_subset_clean['mape_flux'])

                            grouped = mode_subset_clean.groupby(['model_class', 'optimization_method'])['mape_flux'].mean()
                            grouped = grouped.dropna()

                            if len(grouped) > 0:
                                best_idx = grouped.idxmin()
                                # Format target name for single energy modes
                                if mode in ['thermal_only', 'epithermal_only', 'fast_only']:
                                    target_name = f"{mode.replace('_only', '')} flux"
                                else:
                                    target_name = f'flux ({mode})'

                                best_models_data.append({
                                    'Encoding': encoding,
                                    'Target': target_name,
                                    'Best Model': best_idx[0],
                                    'Optimization': best_idx[1],
                                    'Avg Error (%)': grouped.min()
                                })
                else:
                    # Original logic for total flux
                    if 'mape_flux' in flux_subset:
                        # Use absolute values for best model selection
                        flux_subset_abs = flux_subset.copy()
                        flux_subset_abs['mape_flux'] = abs(flux_subset['mape_flux'])
                        grouped = flux_subset_abs.groupby(['model_class', 'optimization_method'])['mape_flux'].mean()
                        if len(grouped) > 0:
                            best_idx = grouped.idxmin()
                            best_models_data.append({
                                'Encoding': encoding,
                                'Target': 'flux',
                                'Best Model': best_idx[0],
                                'Optimization': best_idx[1],
                                'Avg Error (%)': grouped.min()
                            })

        if best_models_data:
            best_models_df = pd.DataFrame(best_models_data)
            best_models_df.to_excel(writer, sheet_name='Best Models by Encoding', index=False)

            # Format best models sheet
            ws_best = writer.sheets['Best Models by Encoding']
            self._format_best_models_sheet(ws_best, best_models_df)

    def _format_worksheet(self, ws, df, has_keff, has_flux, flux_mode='total'):
        """Apply formatting to the main results worksheet"""
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format columns
        for col_idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)

            # Format in_training column - center align T/F values
            if col == 'in_training':
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    # Optionally add color coding
                    if cell.value == 'T':
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                    else:
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red

            # Format percentage/error columns
            elif 'rel_error' in col or 'Error' in col or 'mape' in col.lower():
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'

            # Format scientific notation for flux columns
            elif (('flux' in col or col.startswith('I_')) and
                  ('_real' in col or '_predicted' in col) and
                  'error' not in col):
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00E+00'

            # Format k-eff values
            elif 'keff' in col and 'error' not in col:
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.000000'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    # Handle N/A values
                    cell_value = str(cell.value) if cell.value != 'N/A' else 'N/A'
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass

            # Limit width for readability
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

        # For energy/bin modes, use slightly narrower columns due to many columns
        if flux_mode in ['energy', 'bin']:
            for column in ws.columns:
                column_letter = get_column_letter(column[0].column)
                current_width = ws.column_dimensions[column_letter].width
                ws.column_dimensions[column_letter].width = min(current_width, 18)

    def _format_summary_sheet(self, ws, df):
        """Apply formatting to summary sheet"""
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format percentage columns
        for col_idx, col in enumerate(df.columns, 1):
            if 'Error' in col:
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value is not None:
                        cell.number_format = '0.00'

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _format_best_models_sheet(self, ws, df):
        """Apply formatting to best models sheet"""
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format error column
        if 'Avg Error (%)' in df.columns:
            col_idx = df.columns.get_loc('Avg Error (%)') + 1
            col_letter = get_column_letter(col_idx)
            for row in range(2, len(df) + 2):
                cell = ws[f"{col_letter}{row}"]
                if cell.value is not None:
                    cell.number_format = '0.00'

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
