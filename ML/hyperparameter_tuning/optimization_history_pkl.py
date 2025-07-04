#!/usr/bin/env python3
"""
Optimization History Tracker - Model Files Version

This script extracts optimization results directly from saved model files
and maintains a running history of the best parameters for each model/encoding/optimization
combination in an Excel file.
"""

import os
import re
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import joblib
from datetime import datetime

class OptimizationHistoryTracker:
    def __init__(self):
        self.models_dir = "/root/IntegratedReactorModel/ML/outputs/models"
        self.excel_file = "/root/IntegratedReactorModel/ML/hyperparameter_tuning/optimisation_history.xlsx"

        # Define grouped parameters by model type
        self.svm_params = [
            'C', 'epsilon', 'gamma', 'kernel', 'degree', 'coef0', 'shrinking', 'max_iter', 'tol'
        ]

        self.random_forest_params = [
            'n_estimators', 'max_depth', 'max_features', 'min_samples_split',
            'min_samples_leaf', 'bootstrap', 'max_leaf_nodes', 'max_samples'
        ]

        self.xgboost_params = [
            'n_estimators', 'max_depth', 'learning_rate', 'subsample',
            'colsample_bytree', 'colsample_bylevel', 'reg_alpha', 'reg_lambda',
            'gamma', 'min_child_weight'
        ]

        self.neural_network_params = [
            'hidden_layer_sizes', 'activation', 'solver', 'alpha',
            'learning_rate_init', 'max_iter', 'early_stopping'
        ]

        # Create a single list of all unique parameters
        unique_params = []
        seen = set()

        # Add all parameters from all models, but only once
        for param_list in [self.svm_params, self.random_forest_params,
                          self.xgboost_params, self.neural_network_params]:
            for param in param_list:
                if param not in seen:
                    unique_params.append(param)
                    seen.add(param)

        self.all_params = unique_params

    def scan_model_files(self) -> List[str]:
        """Scan models directory for all model files"""
        models_path = Path(self.models_dir)

        if not models_path.exists():
            print(f"Error: Models directory {self.models_dir} does not exist.")
            return []

        # Get all .pkl files
        model_files = list(models_path.glob("*.pkl"))

        if not model_files:
            print(f"No model files found in {self.models_dir}")
            return []

        print(f"\n📂 Found {len(model_files)} model files in {self.models_dir}")
        return [str(f) for f in sorted(model_files)]

    def parse_model_filename(self, filename: str) -> Optional[Dict]:
        """Parse model filename to extract model type, target, encoding, and optimization method"""
        # Extract just the filename without path
        base_name = os.path.basename(filename)

        # Remove .pkl extension
        name_parts = base_name.replace('.pkl', '')

        # Expected format: model_target_encoding_optimization.pkl
        # e.g., random_forest_flux_physics_three_stage.pkl
        parts = name_parts.split('_')

        if len(parts) < 4:
            print(f"Warning: Unexpected filename format: {base_name}")
            return None

        # Handle model names with underscores (random_forest, neural_network)
        if parts[0] == 'random' and len(parts) > 1 and parts[1] == 'forest':
            model = 'random_forest'
            remaining_parts = parts[2:]
        elif parts[0] == 'neural' and len(parts) > 1 and parts[1] == 'network':
            model = 'neural_network'
            remaining_parts = parts[2:]
        else:
            model = parts[0]
            remaining_parts = parts[1:]

        if len(remaining_parts) < 3:
            print(f"Warning: Not enough parts after model name in: {base_name}")
            return None

        target = remaining_parts[0]

        # Handle optimization methods with underscores (three_stage, one_hot)
        if remaining_parts[-2] == 'three' and remaining_parts[-1] == 'stage':
            optimization = 'three_stage'
            encoding_parts = remaining_parts[1:-2]
        elif remaining_parts[-2] == 'one' and remaining_parts[-1] == 'hot':
            # This is actually encoding, not optimization
            optimization = remaining_parts[-1]
            encoding = 'one_hot'
            encoding_parts = []
        else:
            optimization = remaining_parts[-1]
            encoding_parts = remaining_parts[1:-1]

        # Handle encoding names with underscores
        if encoding_parts:
            if len(encoding_parts) == 2 and encoding_parts[0] == 'one' and encoding_parts[1] == 'hot':
                encoding = 'one_hot'
            else:
                encoding = '_'.join(encoding_parts)
        elif 'encoding' not in locals():
            encoding = encoding_parts[0] if encoding_parts else 'unknown'

        return {
            'filename': base_name,
            'model': model,
            'target': target,
            'encoding': encoding,
            'optimization': optimization
        }

    def extract_model_data(self, model_path: str) -> Optional[Dict]:
        """Extract data from a model pickle file"""
        try:
            # Load the model data
            model_data = joblib.load(model_path)

            # Parse filename for basic info
            file_info = self.parse_model_filename(model_path)
            if not file_info:
                return None

            result = {
                'Model': file_info['model'],
                'Target': file_info['target'],
                'Encoding': file_info['encoding'],
                'Optimisation': file_info['optimization']
            }

            # Check if it's a dictionary with model info
            if isinstance(model_data, dict):
                # Extract flux mode
                if result['Target'] == 'flux':
                    result['Flux Mode'] = model_data.get('flux_mode', 'total')
                else:
                    result['Flux Mode'] = np.nan

                # Extract metrics
                metrics = model_data.get('metrics', {})
                if metrics:
                    result['Test MAPE'] = metrics.get('mape', np.nan)
                    result['Test R²'] = metrics.get('r2', np.nan)
                    result['Test MAE'] = metrics.get('mae', np.nan)
                    result['Test RMSE'] = metrics.get('rmse', np.nan)
                    result['Test MSE'] = metrics.get('mse', np.nan)
                else:
                    # Try alternative metric names
                    result['Test MAPE'] = model_data.get('test_mape', np.nan)
                    result['Test R²'] = model_data.get('test_r2', np.nan)
                    result['Test MAE'] = model_data.get('test_mae', np.nan)
                    result['Test RMSE'] = model_data.get('test_rmse', np.nan)
                    result['Test MSE'] = model_data.get('test_mse', np.nan)

                # Extract parameters
                params = model_data.get('params', {})
                if params:
                    for param_name, param_value in params.items():
                        if param_name in self.all_params:
                            result[param_name] = param_value

                # Try to get training time if available
                result['Training Time (minutes)'] = model_data.get('training_time_minutes', np.nan)

                # If saved_at is available, extract it
                saved_at = model_data.get('saved_at', '')
                if saved_at:
                    result['Last Updated'] = saved_at

            else:
                print(f"Warning: Model file {file_info['filename']} is not in expected dictionary format")
                return None

            return result

        except Exception as e:
            print(f"Error loading model file {model_path}: {e}")
            return None

    def load_existing_history(self) -> pd.DataFrame:
        """Load existing optimization history Excel file from all sheets"""
        base_columns = ['Model', 'Target', 'Flux Mode', 'Encoding', 'Optimisation',
                       'Test MAPE', 'Test R²', 'Test MAE', 'Test RMSE', 'Test MSE',
                       'Training Time (minutes)']

        if os.path.exists(self.excel_file):
            try:
                # Read all sheets
                excel_file = pd.ExcelFile(self.excel_file)
                all_dfs = []

                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(self.excel_file, sheet_name=sheet_name)

                    # Add Target and Flux Mode columns based on sheet name
                    if 'Keff' in sheet_name:
                        df['Target'] = 'keff'
                        df['Flux Mode'] = np.nan
                    elif 'Flux' in sheet_name:
                        df['Target'] = 'flux'
                        # Extract flux mode from sheet name
                        if 'Total' in sheet_name:
                            df['Flux Mode'] = 'total'
                        elif 'Fast' in sheet_name:
                            df['Flux Mode'] = 'fast'
                        elif 'Epithermal' in sheet_name:
                            df['Flux Mode'] = 'epithermal'
                        elif 'Thermal' in sheet_name:
                            df['Flux Mode'] = 'thermal'
                        else:
                            df['Flux Mode'] = 'unknown'

                    all_dfs.append(df)
                    print(f"  Loaded {len(df)} records from sheet: {sheet_name}")

                # Combine all dataframes
                if all_dfs:
                    df = pd.concat(all_dfs, ignore_index=True)
                    print(f"Total records loaded: {len(df)}")

                    # Define the EXACT column order we want
                    column_order = base_columns + self.all_params

                    # Make sure each column appears only once
                    seen = set()
                    unique_columns = []
                    for col in column_order:
                        if col not in seen:
                            seen.add(col)
                            unique_columns.append(col)

                    # Create a new dataframe with ONLY these columns
                    result_df = pd.DataFrame(index=df.index)

                    # Copy data for each column
                    for col in unique_columns:
                        if col in df.columns:
                            result_df[col] = df[col].values
                        else:
                            result_df[col] = np.nan

                    return result_df
                else:
                    # Create empty dataframe
                    return pd.DataFrame(columns=base_columns + self.all_params)

            except Exception as e:
                print(f"Error loading existing Excel file: {e}")
                print("Creating new history file...")

        # Create new DataFrame with base columns + all params
        return pd.DataFrame(columns=base_columns + self.all_params)

    def update_history(self, existing_df: pd.DataFrame, new_results: List[Dict]) -> Tuple[pd.DataFrame, List[Dict]]:
        """Update history with new results, only keeping better MAPE scores"""
        updated_df = existing_df.copy()
        updates_made = 0
        additions_made = 0
        update_details = []

        for result in new_results:
            # Create unique identifier for this combination
            if result['Target'] == 'flux':
                # For flux targets, include flux mode in the identifier
                mask = (
                    (updated_df['Model'] == result['Model']) &
                    (updated_df['Target'] == result['Target']) &
                    (updated_df['Encoding'] == result['Encoding']) &
                    (updated_df['Optimisation'] == result['Optimisation']) &
                    (updated_df['Flux Mode'] == result.get('Flux Mode', 'total'))
                )
                flux_mode = result.get('Flux Mode', 'total')
                identifier_str = f"{result['Model']}-{flux_mode} flux-{result['Encoding']}-{result['Optimisation']}"
            else:
                # For keff targets, flux mode doesn't matter
                mask = (
                    (updated_df['Model'] == result['Model']) &
                    (updated_df['Target'] == result['Target']) &
                    (updated_df['Encoding'] == result['Encoding']) &
                    (updated_df['Optimisation'] == result['Optimisation'])
                )
                identifier_str = f"{result['Model']}-keff-{result['Encoding']}-{result['Optimisation']}"

            existing_records = updated_df[mask]

            if len(existing_records) > 0:
                # Record exists, check if new result is better
                existing_idx = existing_records.index[0]
                existing_mape = updated_df.loc[existing_idx, 'Test MAPE']
                new_mape = result.get('Test MAPE', np.nan)

                # Compare based on MAPE (lower is better)
                if pd.notna(new_mape) and (pd.isna(existing_mape) or new_mape < existing_mape):
                    # Update with better result
                    for col, val in result.items():
                        if col in updated_df.columns:
                            if pd.notna(val):
                                updated_df.at[existing_idx, col] = val

                    updates_made += 1
                    update_info = {
                        'action': 'updated',
                        'identifier': identifier_str,
                        'row': existing_idx + 2,  # +2 for Excel row (1-indexed + header)
                        'old_mape': existing_mape,
                        'new_mape': new_mape
                    }
                    update_details.append(update_info)

                    if pd.isna(existing_mape):
                        print(f"✅ Updated {identifier_str}: MAPE N/A → {new_mape:.4f}%")
                    else:
                        print(f"✅ Updated {identifier_str}: MAPE {existing_mape:.4f}% → {new_mape:.4f}%")
                else:
                    if pd.notna(new_mape) and pd.notna(existing_mape):
                        print(f"⏭️  Skipped {identifier_str}: New MAPE ({new_mape:.4f}%) not better than existing ({existing_mape:.4f}%)")
            else:
                # New combination, add it
                new_row = {col: np.nan for col in updated_df.columns}
                for col, val in result.items():
                    if col in new_row:
                        new_row[col] = val

                new_row_df = pd.DataFrame([new_row])
                updated_df = pd.concat([updated_df, new_row_df], ignore_index=True)
                additions_made += 1

                update_info = {
                    'action': 'added',
                    'identifier': identifier_str,
                    'row': len(updated_df) + 1,
                    'new_mape': result.get('Test MAPE', np.nan)
                }
                update_details.append(update_info)

                mape_str = f"{result.get('Test MAPE', 'N/A'):.4f}%" if pd.notna(result.get('Test MAPE', np.nan)) else "N/A"
                print(f"➕ Added new: {identifier_str} (MAPE: {mape_str})")

        print(f"\n📊 Summary: {updates_made} records updated, {additions_made} records added")
        return updated_df, update_details

    def save_history(self, df: pd.DataFrame, update_details: List[Dict]):
        """Save updated history to Excel file with proper formatting"""
        try:
            # Ensure directory exists
            os.makedirs(os.path.dirname(self.excel_file), exist_ok=True)

            # Sort data
            df_sorted = df.sort_values(['Model', 'Target', 'Flux Mode', 'Encoding', 'Optimisation'],
                                     na_position='last').reset_index(drop=True)

            # Split by target type
            flux_df = df_sorted[df_sorted['Target'] == 'flux'].copy()
            keff_df = df_sorted[df_sorted['Target'] == 'keff'].copy()

            # Create Excel writer
            with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                sheets_created = []

                # Create sheets for each flux mode
                if not flux_df.empty:
                    flux_modes = flux_df['Flux Mode'].dropna().unique()

                    for flux_mode in flux_modes:
                        flux_mode_df = flux_df[flux_df['Flux Mode'] == flux_mode].copy()
                        if not flux_mode_df.empty:
                            # Remove only Target and Flux Mode columns, keep all others including parameters
                            cols_to_keep = [col for col in flux_mode_df.columns if col not in ['Target', 'Flux Mode']]
                            flux_mode_df = flux_mode_df[cols_to_keep]
                            sheet_name = f'{flux_mode.title()} Flux Results'
                            flux_mode_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            self._format_sheet(writer.sheets[sheet_name], flux_mode_df)
                            sheets_created.append(f"{sheet_name}: {len(flux_mode_df)} records")

                # Create keff sheet
                if not keff_df.empty:
                    # Remove only Target and Flux Mode columns, keep all others including parameters
                    cols_to_keep = [col for col in keff_df.columns if col not in ['Target', 'Flux Mode']]
                    keff_df = keff_df[cols_to_keep]
                    keff_df.to_excel(writer, sheet_name='Keff Results', index=False)
                    self._format_sheet(writer.sheets['Keff Results'], keff_df)
                    sheets_created.append(f"Keff Results: {len(keff_df)} records")

            print(f"\n💾 Optimization history saved to: {self.excel_file}")
            for sheet_info in sheets_created:
                print(f"  📄 {sheet_info}")

            # Print which specific rows were updated
            if update_details:
                print("\n📝 Update details:")
                for detail in update_details:
                    if detail['action'] == 'updated':
                        if pd.isna(detail['old_mape']):
                            print(f"  - Row {detail['row']}: {detail['identifier']} updated (MAPE: N/A → {detail['new_mape']:.4f}%)")
                        else:
                            print(f"  - Row {detail['row']}: {detail['identifier']} updated (MAPE: {detail['old_mape']:.4f}% → {detail['new_mape']:.4f}%)")
                    else:
                        mape_str = f"{detail['new_mape']:.4f}%" if pd.notna(detail['new_mape']) else "N/A"
                        print(f"  - Row {detail['row']}: {detail['identifier']} added (MAPE: {mape_str})")

        except Exception as e:
            print(f"Error saving Excel file: {e}")

    def _format_sheet(self, ws, df):
        """Format Excel sheet with colors and styling"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Define colors
        colors = {
            'xgboost': PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # Light blue
            'random_forest': PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid"),  # Light orange
            'svm': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light green
            'neural_network': PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid"),  # Light purple
            'param_header': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),  # Light grey
            'param_cell': PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Very light grey
        }

        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Format headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Update parameter headers with expanded names
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in self.all_params:
                col_letter = get_column_letter(col_idx)
                expanded_name = self._expand_parameter_name(col_name)
                ws[f"{col_letter}1"].value = expanded_name
                ws[f"{col_letter}1"].fill = colors['param_header']
                ws[f"{col_letter}1"].font = Font(bold=True, color="333333")

        # Format data rows
        for row_idx in range(2, len(df) + 2):
            model = df.iloc[row_idx - 2]['Model'].lower()
            model_color = colors.get(model, PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))

            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                cell = ws[f"{col_letter}{row_idx}"]

                # Apply appropriate coloring
                if col_name not in self.all_params:
                    cell.fill = model_color
                else:
                    cell.fill = colors['param_cell']

                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Format numeric values
                if col_name in ['Test MAPE', 'Test R²', 'Test MAE', 'Test RMSE', 'Test MSE']:
                    if pd.notna(cell.value) and isinstance(cell.value, (int, float)):
                        if 'R²' in col_name:
                            cell.number_format = '0.0000'
                        elif 'MAPE' in col_name:
                            cell.number_format = '0.0000'
                        else:
                            cell.number_format = '0.00E+00'
                elif col_name == 'Training Time (minutes)':
                    if pd.notna(cell.value) and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, 10), 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _expand_parameter_name(self, param_name):
        """Expand parameter names to be more descriptive"""
        expansions = {
            'n_estimators': 'Number of Estimators',
            'max_depth': 'Maximum Depth',
            'learning_rate': 'Learning Rate',
            'subsample': 'Subsample Ratio',
            'colsample_bytree': 'Column Sample by Tree',
            'colsample_bylevel': 'Column Sample by Level',
            'reg_alpha': 'L1 Regularization (Alpha)',
            'reg_lambda': 'L2 Regularization (Lambda)',
            'gamma': 'Gamma (Min Split Loss)',
            'min_child_weight': 'Min Child Weight',
            'max_features': 'Max Features',
            'min_samples_split': 'Min Samples Split',
            'min_samples_leaf': 'Min Samples Leaf',
            'bootstrap': 'Bootstrap Sampling',
            'max_leaf_nodes': 'Max Leaf Nodes',
            'max_samples': 'Max Samples Ratio',
            'C': 'Regularization Parameter (C)',
            'epsilon': 'Epsilon (SVR)',
            'kernel': 'Kernel Type',
            'degree': 'Polynomial Degree',
            'coef0': 'Independent Term (coef0)',
            'tol': 'Tolerance',
            'hidden_layer_sizes': 'Hidden Layer Sizes',
            'activation': 'Activation Function',
            'solver': 'Solver Algorithm',
            'alpha': 'L2 Regularization',
            'learning_rate_init': 'Initial Learning Rate',
            'max_iter': 'Max Iterations',
            'early_stopping': 'Early Stopping',
            'shrinking': 'Use Shrinking Heuristic'
        }
        return expansions.get(param_name, param_name)

    def run(self):
        """Main execution function"""
        print("="*60)
        print("OPTIMIZATION HISTORY TRACKER - Model Files Version")
        print("="*60)

        # Scan for model files
        model_files = self.scan_model_files()
        if not model_files:
            return

        # Extract data from each model file
        print(f"\n📖 Extracting data from model files...")
        results = []

        for i, model_path in enumerate(model_files, 1):
            print(f"\r  Processing {i}/{len(model_files)}: {os.path.basename(model_path)}", end='', flush=True)

            model_data = self.extract_model_data(model_path)
            if model_data:
                results.append(model_data)

        print(f"\n✅ Successfully extracted data from {len(results)} model files")

        if not results:
            print("❌ No valid model data found.")
            return

        # Load existing history
        print(f"\n📂 Loading existing history from: {self.excel_file}")
        existing_df = self.load_existing_history()

        # Update history
        print(f"\n🔄 Updating history...")
        updated_df, update_details = self.update_history(existing_df, results)

        # Save updated history
        self.save_history(updated_df, update_details)

def main():
    tracker = OptimizationHistoryTracker()
    tracker.run()

if __name__ == "__main__":
    main()
