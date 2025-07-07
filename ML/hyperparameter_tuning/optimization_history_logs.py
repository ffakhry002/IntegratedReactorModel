#!/usr/bin/env python3
"""
Optimization History Tracker

This script parses ML training log files to extract optimization results
and maintains a running history of the best parameters for each model/encoding/optimization
combination in an Excel file.
"""

import os
import re
import pandas as pd
import ast
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import numpy as np

class OptimizationHistoryTracker:
    def __init__(self):
        """Initialize the optimization history tracker.

        Parameters
        ----------
        None

        Returns
        -------
        None
        """
        self.log_dir = "ML/outputs/logs"
        self.excel_file = "ML/hyperparameter_tuning/optimisation_history_logs.xlsx"

        # Define grouped parameters by model type
        # Note: Some parameters appear in multiple models (e.g., n_estimators, max_depth, gamma)
        # We'll only include each unique parameter once in the final list

        self.svm_params = [
            'C', 'epsilon', 'gamma', 'kernel', 'degree', 'coef0', 'shrinking', 'max_iter', 'tol'
        ]

        self.random_forest_params = [
            'n_estimators', 'max_depth', 'max_features', 'min_samples_split',
            'min_samples_leaf', 'bootstrap', 'max_leaf_nodes', 'max_samples'
        ]

        self.xgboost_params = [
            'n_estimators', 'max_depth', 'learning_rate', 'subsample',
            'colsample_bytree', 'colsample_bylevel', 'reg_alpha', 'reg_lambda',
            'gamma', 'min_child_weight'
        ]

        self.neural_network_params = [
            'hidden_layer_sizes', 'activation', 'solver', 'alpha',
            'learning_rate_init', 'max_iter', 'early_stopping'
        ]

        # Create a single list of all unique parameters
        # This ensures each parameter appears only once, regardless of how many models use it
        unique_params = []
        seen = set()

        # Add all parameters from all models, but only once
        for param_list in [self.svm_params, self.random_forest_params,
                          self.xgboost_params, self.neural_network_params]:
            for param in param_list:
                if param not in seen:
                    unique_params.append(param)
                    seen.add(param)

        self.all_params = unique_params

        # Create a single list of all unique parameters
        # This ensures each parameter appears only once, regardless of how many models use it
        all_params_list = []
        seen_params = set()

        # Add SVM params first
        for param in self.svm_params:
            if param not in seen_params:
                all_params_list.append(param)
                seen_params.add(param)

        # Add Random Forest params (skip duplicates like n_estimators, max_depth)
        for param in self.random_forest_params:
            if param not in seen_params:
                all_params_list.append(param)
                seen_params.add(param)

        # Add XGBoost params (skip duplicates)
        for param in self.xgboost_params:
            if param not in seen_params:
                all_params_list.append(param)
                seen_params.add(param)

        # Add Neural Network params (skip duplicates)
        for param in self.neural_network_params:
            if param not in seen_params:
                all_params_list.append(param)
                seen_params.add(param)

        self.all_params = all_params_list

    def select_log_file(self) -> str:
        """Interactive log file selection"""
        log_path = Path(self.log_dir)

        if not log_path.exists():
            print(f"Error: Log directory {self.log_dir} does not exist.")
            return None

        log_files = list(log_path.glob("*.log"))
        if not log_files:
            print(f"No log files found in {self.log_dir}")
            return None

        print(f"\nAvailable log files in {self.log_dir}:")
        print("-" * 50)

        for i, log_file in enumerate(log_files, 1):
            file_size = log_file.stat().st_size / (1024 * 1024)  # MB
            print(f"{i:2d}. {log_file.name} ({file_size:.1f} MB)")

        print(f"{len(log_files) + 1:2d}. Custom path")
        print("-" * 50)

        while True:
            try:
                choice = input(f"Select log file (1-{len(log_files) + 1}): ").strip()
                if choice.isdigit():
                    choice_num = int(choice)
                    if 1 <= choice_num <= len(log_files):
                        selected_file = log_files[choice_num - 1]
                        print(f"Selected: {selected_file}")
                        return str(selected_file)
                    elif choice_num == len(log_files) + 1:
                        custom_path = input("Enter custom log file path: ").strip()
                        if os.path.exists(custom_path):
                            return custom_path
                        else:
                            print("File not found. Please try again.")
                    else:
                        print(f"Please enter a number between 1 and {len(log_files) + 1}")
                else:
                    print("Please enter a valid number")
            except (ValueError, KeyboardInterrupt):
                print("\nOperation cancelled.")
                return None

    def parse_log_file(self, log_file_path: str) -> List[Dict]:
        """Parse log file to extract optimization results"""
        results = []

        try:
            with open(log_file_path, 'r', encoding='utf-8') as f:
                content = f.read()
        except Exception as e:
            print(f"Error reading log file: {e}")
            return results

        # Split content into job sections
        job_sections = self._split_into_job_sections(content)
        print(f"Found {len(job_sections)} job sections in log file")

        for i, section in enumerate(job_sections, 1):
            result = self._parse_job_section(section)
            if result:
                results.append(result)
                mape_str = f"{result.get('Test MAPE', 'N/A'):.2f}%" if isinstance(result.get('Test MAPE'), (int, float)) else "N/A"
                print(f"  ✓ Parsed job {i}/{len(job_sections)}: {result['Model']}-{result['Encoding']}-{result['Optimisation']} (MAPE: {mape_str})")
            else:
                print(f"  ✗ Failed to parse job {i}/{len(job_sections)}")

        return results

    def _split_into_job_sections(self, content: str) -> List[str]:
        """Split log content into individual job sections"""
        sections = []

        # Find all job starts
        job_start_pattern = r'Job \d+/\d+:'
        job_starts = list(re.finditer(job_start_pattern, content))

        for i, match in enumerate(job_starts):
            start_pos = match.start()
            end_pos = job_starts[i + 1].start() if i + 1 < len(job_starts) else len(content)
            section = content[start_pos:end_pos]

            # Only include sections that have encoding and optimization info
            if 'Encoding:' in section and 'Optimization:' in section:
                sections.append(section)

        return sections

    def _parse_job_section(self, section: str) -> Optional[Dict]:
        """Parse a single job section to extract results"""
        result = {}

        # Extract basic info
        job_line_match = re.search(r'Job \d+/\d+: (\w+) for (\w+)', section)
        encoding_match = re.search(r'Encoding: (\w+)', section)
        optimization_match = re.search(r'Optimization: (\w+)', section)

        if not (job_line_match and encoding_match and optimization_match):
            return None

        result['Model'] = job_line_match.group(1)
        result['Target'] = job_line_match.group(2)
        result['Encoding'] = encoding_match.group(1)
        result['Optimisation'] = optimization_match.group(1)

        # Extract flux mode for flux targets
        if result['Target'] == 'flux':
            flux_mode_match = re.search(r'Flux mode: (\w+)', section)
            result['Flux Mode'] = flux_mode_match.group(1) if flux_mode_match else 'unknown'
        else:
            result['Flux Mode'] = np.nan

        # Extract metrics
        mape_match = re.search(r'Test MAPE: ([\d.]+)%', section)
        if mape_match:
            result['Test MAPE'] = float(mape_match.group(1))
        else:
            result['Test MAPE'] = np.nan

        # Extract other metrics
        metric_patterns = {
            'Test R²': r'Test R²: ([\d.-]+)',
            'Test MAE': r'Test MAE: ([\d.e+-]+)',
            'Test RMSE': r'Test RMSE: ([\d.e+-]+)',
            'Test MSE': r'Test MSE: ([\d.e+-]+)',
        }

        for metric_name, pattern in metric_patterns.items():
            match = re.search(pattern, section)
            if match:
                result[metric_name] = float(match.group(1))

        # Extract training time
        time_match = re.search(r'(?:Optimization took|Total time for \w+ \w+:) ([\d.]+) minutes', section)
        if time_match:
            result['Training Time (minutes)'] = float(time_match.group(1))

        # Extract parameters based on optimization type
        if result['Optimisation'] == 'none':
            # For 'none' optimization, extract default parameters
            params = self._extract_default_parameters(section, result['Model'])
            result.update(params)
        elif result['Optimisation'] in ['optuna', 'three_stage']:
            # Look for different parameter formats
            params = self._extract_parameters_from_section(section, result['Model'])
            result.update(params)

        return result

    def _extract_default_parameters(self, section: str, model_type: str) -> Dict:
        """Extract default parameters for 'none' optimization method"""
        params = {}

        # Look for "Using default parameters:" line
        default_params_match = re.search(r'Using default parameters: ({.*?})', section, re.DOTALL)
        if default_params_match:
            try:
                params_dict = ast.literal_eval(default_params_match.group(1))
                return params_dict
            except Exception as e:
                print(f"Warning: Could not parse default parameters: {e}")

        return params

    def _extract_parameters_from_section(self, section: str, model_type: str) -> Dict:
        """Extract parameters from optimization section, handling various formats"""
        params = {}

        # Format 1: Best parameters found: {...}
        params_match = re.search(r'Best parameters found: ({.*?})', section, re.DOTALL)
        if params_match:
            try:
                params_dict = ast.literal_eval(params_match.group(1))
                return params_dict
            except:
                pass

        # Format 2: Three-stage optimization patterns
        # Look for "Best parameters (cleaned):" or "Best parameters:" followed by individual parameters
        three_stage_patterns = [
            r'Best parameters \(cleaned\):(.*?)(?=\n[A-Z]|Optimization complete|Training final model|$)',
            r'Best parameters:(.*?)(?=\n[A-Z]|Optimization complete|Training final model|$)',
        ]

        for pattern in three_stage_patterns:
            params_match = re.search(pattern, section, re.DOTALL)
            if params_match:
                param_section = params_match.group(1)
                # Extract individual parameters from the section
                extracted_params = self._extract_individual_parameters(param_section)
                if extracted_params:
                    return extracted_params

        # Format 3: Final best parameters listed individually
        # Look for "Optimal parameters" or "Optimized parameters" section
        optimal_section_match = re.search(r'(?:Optimal|Optimized|Final best) parameters.*?(?=Fixed parameters|Optimization complete|Training final model|$)', section, re.DOTALL)
        if optimal_section_match:
            param_section = optimal_section_match.group(0)
            extracted_params = self._extract_individual_parameters(param_section)
            if extracted_params:
                return extracted_params

        # If no parameters found yet, also check for the section after "Final best MAPE"
        if not params:
            # Try to find parameters after optimization complete message
            final_section_match = re.search(r'Final best MAPE:.*?(?:$|\n\n)', section, re.DOTALL)
            if final_section_match:
                # Look in the entire remaining section after Final best MAPE
                remaining_section = section[final_section_match.start():]
                return self._extract_parameters_from_section(remaining_section, model_type)

        return params

    def _extract_individual_parameters(self, param_section: str) -> Dict:
        """Extract individual parameters from a section with lines like '- param: value'"""
        params = {}

        # Extract individual parameters - expanded list to catch all possible parameters
        param_patterns = {
            # SVM parameters
            'C': r'- C: ([\d.e+-]+)',
            'epsilon': r'- epsilon: ([\d.e+-]+)',
            'gamma': r'- gamma: ([\d.e+-]+)',
            'kernel': r'- kernel: (\w+)',
            'shrinking': r'- shrinking: (\w+)',
            'max_iter': r'- max_iter: ([-\d]+)',
            'degree': r'- degree: (\d+)',
            'coef0': r'- coef0: ([\d.e+-]+)',
            'tol': r'- tol: ([\d.e+-]+)',
            # XGBoost parameters
            'n_estimators': r'- n_estimators: (\d+)',
            'max_depth': r'- max_depth: (\d+)',
            'learning_rate': r'- learning_rate: ([\d.e+-]+)',
            'subsample': r'- subsample: ([\d.e+-]+)',
            'colsample_bytree': r'- colsample_bytree: ([\d.e+-]+)',
            'colsample_bylevel': r'- colsample_bylevel: ([\d.e+-]+)',
            'reg_alpha': r'- reg_alpha: ([\d.e+-]+)',
            'reg_lambda': r'- reg_lambda: ([\d.e+-]+)',
            'min_child_weight': r'- min_child_weight: (\d+)',
            # Random Forest parameters
            'max_features': r'- max_features: ([\w.]+)',
            'min_samples_split': r'- min_samples_split: (\d+)',
            'min_samples_leaf': r'- min_samples_leaf: (\d+)',
            'bootstrap': r'- bootstrap: (\w+)',
            'max_leaf_nodes': r'- max_leaf_nodes: (\d+)',
            'max_samples': r'- max_samples: ([\d.e+-]+)',
            # Neural Network parameters
            'hidden_layer_sizes': r'- hidden_layer_sizes: ([\(\)\d,\s]+)',
            'activation': r'- activation: (\w+)',
            'solver': r'- solver: (\w+)',
            'alpha': r'- alpha: ([\d.e+-]+)',
            'learning_rate_init': r'- learning_rate_init: ([\d.e+-]+)',
            'early_stopping': r'- early_stopping: (\w+)',
        }

        for param_name, pattern in param_patterns.items():
            match = re.search(pattern, param_section)
            if match:
                value = match.group(1)
                # Convert to appropriate type
                if value in ['True', 'False']:
                    params[param_name] = value == 'True'
                elif value.replace('.', '').replace('-', '').replace('e', '').replace('+', '').isdigit():
                    if '.' in value or 'e' in value or 'E' in value:
                        params[param_name] = float(value)
                    else:
                        params[param_name] = int(value)
                else:
                    params[param_name] = value
                print(f"      Extracted {param_name}: {params[param_name]}")

        return params

    def load_existing_history(self) -> pd.DataFrame:
        """Load existing optimization history Excel file from all sheets"""
        base_columns = ['Model', 'Target', 'Flux Mode', 'Encoding', 'Optimisation',
                       'Test MAPE', 'Test R²', 'Test MAE', 'Test RMSE', 'Test MSE',
                       'Training Time (minutes)']

        if os.path.exists(self.excel_file):
            try:
                # Read all sheets
                excel_file = pd.ExcelFile(self.excel_file)
                all_dfs = []

                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(self.excel_file, sheet_name=sheet_name)

                    # Add Target and Flux Mode columns based on sheet name
                    if 'Keff' in sheet_name:
                        df['Target'] = 'keff'
                        df['Flux Mode'] = np.nan
                    elif 'Flux' in sheet_name:
                        df['Target'] = 'flux'
                        # Extract flux mode from sheet name
                        if 'Total' in sheet_name:
                            df['Flux Mode'] = 'total'
                        elif 'Fast' in sheet_name:
                            df['Flux Mode'] = 'fast'
                        elif 'Epithermal' in sheet_name:
                            df['Flux Mode'] = 'epithermal'
                        elif 'Thermal' in sheet_name:
                            df['Flux Mode'] = 'thermal'
                        else:
                            df['Flux Mode'] = 'unknown'

                    all_dfs.append(df)
                    print(f"  Loaded {len(df)} records from sheet: {sheet_name}")

                # Combine all dataframes
                if all_dfs:
                    df = pd.concat(all_dfs, ignore_index=True)
                    print(f"Total records loaded: {len(df)}")

                    # Define the EXACT column order we want
                    column_order = base_columns + self.all_params

                    # Make sure each column appears only once
                    seen = set()
                    unique_columns = []
                    for col in column_order:
                        if col not in seen:
                            seen.add(col)
                            unique_columns.append(col)

                    # Create a new dataframe with ONLY these columns
                    result_df = pd.DataFrame(index=df.index)

                    # First, create a mapping of Excel column names to our standard names
                    # This handles cases where Excel might have modified column names
                    excel_to_standard = {}
                    for excel_col in df.columns:
                        # Try exact match first
                        if excel_col in unique_columns:
                            excel_to_standard[excel_col] = excel_col
                        else:
                            # Try to match parameter names that might have been expanded
                            for standard_col in unique_columns:
                                expanded = self._expand_parameter_name(standard_col) if standard_col in self.all_params else standard_col
                                if excel_col == expanded or excel_col.lower() == standard_col.lower():
                                    excel_to_standard[excel_col] = standard_col
                                    break

                    # Copy data for each column
                    for col in unique_columns:
                        # Check if this column exists in Excel (possibly with different name)
                        excel_col = None
                        for excel_name, standard_name in excel_to_standard.items():
                            if standard_name == col:
                                excel_col = excel_name
                                break

                        if excel_col and excel_col in df.columns:
                            result_df[col] = df[excel_col].values
                        else:
                            result_df[col] = np.nan

                    # Debug: Check if we're preserving parameter values
                    param_cols_with_data = []
                    for col in self.all_params:
                        if col in result_df.columns and result_df[col].notna().any():
                            param_cols_with_data.append(col)
                    if param_cols_with_data:
                        print(f"  Preserved parameter data for: {', '.join(param_cols_with_data)}")

                    return result_df
                else:
                    # Create empty dataframe
                    return pd.DataFrame(columns=base_columns + self.all_params)

            except Exception as e:
                print(f"Error loading existing Excel file: {e}")
                import traceback
                traceback.print_exc()
                print("Creating new history file...")

        # Create new DataFrame with base columns + all params
        return pd.DataFrame(columns=base_columns + self.all_params)

    def update_history(self, existing_df: pd.DataFrame, new_results: List[Dict]) -> Tuple[pd.DataFrame, List[Dict]]:
        """Update history with new results, only keeping better MAPE scores"""
        updated_df = existing_df.copy()
        updates_made = 0
        additions_made = 0
        update_details = []

        for result in new_results:
            # Create unique identifier for this combination
            if result['Target'] == 'flux':
                # For flux targets, include flux mode in the identifier
                mask = (
                    (updated_df['Model'] == result['Model']) &
                    (updated_df['Target'] == result['Target']) &
                    (updated_df['Encoding'] == result['Encoding']) &
                    (updated_df['Optimisation'] == result['Optimisation']) &
                    (updated_df['Flux Mode'] == result['Flux Mode'])
                )
                identifier_str = f"{result['Model']}-{result['Flux Mode']} flux-{result['Encoding']}-{result['Optimisation']}"
            else:
                # For keff targets, flux mode doesn't matter
                mask = (
                    (updated_df['Model'] == result['Model']) &
                    (updated_df['Target'] == result['Target']) &
                    (updated_df['Encoding'] == result['Encoding']) &
                    (updated_df['Optimisation'] == result['Optimisation'])
                )
                identifier_str = f"{result['Model']}-keff-{result['Encoding']}-{result['Optimisation']}"

            existing_records = updated_df[mask]

            if len(existing_records) > 0:
                # Record exists, check if new result is better
                existing_idx = existing_records.index[0]
                existing_mape = updated_df.loc[existing_idx, 'Test MAPE']
                new_mape = result.get('Test MAPE', np.nan)

                # Compare based on MAPE (lower is better)
                if pd.notna(new_mape) and (pd.isna(existing_mape) or new_mape < existing_mape):
                    # Update with better result, but preserve existing values where we don't have new ones
                    for col, val in result.items():
                        if col in updated_df.columns:
                            # Only update if we have a new value (not NaN)
                            if pd.notna(val):
                                # Handle dtype compatibility
                                current_dtype = updated_df[col].dtype
                                try:
                                    # For object columns, we can assign directly
                                    if current_dtype == 'object' or pd.api.types.is_object_dtype(current_dtype):
                                        updated_df.at[existing_idx, col] = val
                                    else:
                                        # For numeric columns, try to convert if possible
                                        if isinstance(val, (int, float)):
                                            updated_df.at[existing_idx, col] = val
                                        else:
                                            # If we have a string value for a numeric column,
                                            # we need to convert the column to object first
                                            updated_df[col] = updated_df[col].astype('object')
                                            updated_df.at[existing_idx, col] = val
                                except Exception as e:
                                    print(f"Warning: Could not update {col} with value {val}: {e}")
                            elif col in ['Test MAPE', 'Test R²', 'Test MAE', 'Test RMSE', 'Test MSE', 'Training Time (minutes)']:
                                # For metrics, we can update with NaN if that's what we got
                                updated_df.at[existing_idx, col] = val
                            # For parameters, if the new value is NaN, keep the existing value
                            # (This is implicit - we just don't update it)

                    updates_made += 1
                    update_info = {
                        'action': 'updated',
                        'identifier': identifier_str,
                        'row': existing_idx + 2,  # +2 for Excel row (1-indexed + header)
                        'old_mape': existing_mape,
                        'new_mape': new_mape
                    }
                    update_details.append(update_info)

                    if pd.isna(existing_mape):
                        print(f"Updated {identifier_str}: MAPE N/A → {new_mape:.2f}%")
                    else:
                        print(f"Updated {identifier_str}: MAPE {existing_mape:.2f}% → {new_mape:.2f}%")
                else:
                    if pd.notna(new_mape) and pd.notna(existing_mape):
                        print(f"Skipped {identifier_str}: New MAPE ({new_mape:.2f}%) not better than existing ({existing_mape:.2f}%)")
                    else:
                        print(f"Skipped {identifier_str}: Invalid MAPE comparison")
            else:
                # New combination, add it
                # Start with all existing columns filled with NaN
                new_row = {col: np.nan for col in updated_df.columns}
                # Then update with values from the result
                for col, val in result.items():
                    if col in new_row:
                        new_row[col] = val

                # Create new row DataFrame ensuring no duplicate columns
                new_row_df = pd.DataFrame([new_row])

                # Concatenate the new row
                updated_df = pd.concat([updated_df, new_row_df], ignore_index=True)
                additions_made += 1

                update_info = {
                    'action': 'added',
                    'identifier': identifier_str,
                    'row': len(updated_df) + 1,  # +1 for Excel row (1-indexed)
                    'new_mape': result.get('Test MAPE', np.nan)
                }
                update_details.append(update_info)

                mape_str = f"{result.get('Test MAPE', 'N/A'):.2f}%" if pd.notna(result.get('Test MAPE', np.nan)) else "N/A"
                print(f"➕ Added new: {identifier_str} (MAPE: {mape_str})")

        print(f"\nSummary: {updates_made} records updated, {additions_made} records added")
        return updated_df, update_details

    def save_history(self, df: pd.DataFrame, update_details: List[Dict]):
        """Save updated history to Excel file with proper formatting"""
        try:
            # Ensure directory exists
            os.makedirs(os.path.dirname(self.excel_file), exist_ok=True)

            # Sort data
            df_sorted = df.sort_values(['Model', 'Target', 'Flux Mode', 'Encoding', 'Optimisation'],
                                     na_position='last').reset_index(drop=True)

            # Split by target type
            flux_df = df_sorted[df_sorted['Target'] == 'flux'].copy()
            keff_df = df_sorted[df_sorted['Target'] == 'keff'].copy()

            # Create Excel writer
            with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                sheets_created = []

                # Create sheets for each flux mode
                if not flux_df.empty:
                    flux_modes = flux_df['Flux Mode'].dropna().unique()

                    for flux_mode in flux_modes:
                        flux_mode_df = flux_df[flux_df['Flux Mode'] == flux_mode].copy()
                        if not flux_mode_df.empty:
                            # Remove only Target and Flux Mode columns, keep all others including parameters
                            cols_to_keep = [col for col in flux_mode_df.columns if col not in ['Target', 'Flux Mode']]
                            flux_mode_df = flux_mode_df[cols_to_keep]
                            sheet_name = f'{flux_mode.title()} Flux Results'
                            flux_mode_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            self._format_sheet(writer.sheets[sheet_name], flux_mode_df)
                            sheets_created.append(f"{sheet_name}: {len(flux_mode_df)} records")

                # Create keff sheet
                if not keff_df.empty:
                    # Remove only Target and Flux Mode columns, keep all others including parameters
                    cols_to_keep = [col for col in keff_df.columns if col not in ['Target', 'Flux Mode']]
                    keff_df = keff_df[cols_to_keep]
                    keff_df.to_excel(writer, sheet_name='Keff Results', index=False)
                    self._format_sheet(writer.sheets['Keff Results'], keff_df)
                    sheets_created.append(f"Keff Results: {len(keff_df)} records")

            print(f"\nOptimization history saved to: {self.excel_file}")
            for sheet_info in sheets_created:
                print(f"  {sheet_info}")

            # Print which specific rows were updated
            if update_details:
                print("\nUpdate details:")
                for detail in update_details:
                    if detail['action'] == 'updated':
                        if pd.isna(detail['old_mape']):
                            print(f"  - Row {detail['row']}: {detail['identifier']} updated (MAPE: N/A → {detail['new_mape']:.2f}%)")
                        else:
                            print(f"  - Row {detail['row']}: {detail['identifier']} updated (MAPE: {detail['old_mape']:.2f}% → {detail['new_mape']:.2f}%)")
                    else:
                        mape_str = f"{detail['new_mape']:.2f}%" if pd.notna(detail['new_mape']) else "N/A"
                        print(f"  - Row {detail['row']}: {detail['identifier']} added (MAPE: {mape_str})")

        except Exception as e:
            print(f"Error saving Excel file: {e}")

    def _format_sheet(self, ws, df):
        """Format Excel sheet with colors and styling"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Define colors
        colors = {
            'xgboost': PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # Light blue
            'random_forest': PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid"),  # Light orange
            'svm': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light green
            'neural_network': PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid"),  # Light purple
            'param_header': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),  # Light grey
            'param_cell': PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Very light grey
        }

        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Format headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Update parameter headers with expanded names
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in self.all_params:
                col_letter = get_column_letter(col_idx)
                expanded_name = self._expand_parameter_name(col_name)
                ws[f"{col_letter}1"].value = expanded_name
                ws[f"{col_letter}1"].fill = colors['param_header']
                ws[f"{col_letter}1"].font = Font(bold=True, color="333333")

        # Format data rows
        for row_idx in range(2, len(df) + 2):
            model = df.iloc[row_idx - 2]['Model'].lower()
            model_color = colors.get(model, PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))

            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                cell = ws[f"{col_letter}{row_idx}"]

                # Apply appropriate coloring
                if col_name not in self.all_params:
                    cell.fill = model_color
                else:
                    cell.fill = colors['param_cell']

                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Format numeric values
                if col_name in ['Test MAPE', 'Test R²', 'Test MAE', 'Test RMSE', 'Test MSE']:
                    if pd.notna(cell.value) and isinstance(cell.value, (int, float)):
                        if 'R²' in col_name:
                            cell.number_format = '0.0000'
                        elif 'MAPE' in col_name:
                            cell.number_format = '0.00'
                        else:
                            cell.number_format = '0.00E+00'
                elif col_name == 'Training Time (minutes)':
                    if pd.notna(cell.value) and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, 10), 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _expand_parameter_name(self, param_name):
        """Expand parameter names to be more descriptive"""
        expansions = {
            'n_estimators': 'Number of Estimators',
            'max_depth': 'Maximum Depth',
            'learning_rate': 'Learning Rate',
            'subsample': 'Subsample Ratio',
            'colsample_bytree': 'Column Sample by Tree',
            'colsample_bylevel': 'Column Sample by Level',
            'reg_alpha': 'L1 Regularization (Alpha)',
            'reg_lambda': 'L2 Regularization (Lambda)',
            'gamma': 'Gamma (Min Split Loss)',
            'min_child_weight': 'Min Child Weight',
            'max_features': 'Max Features',
            'min_samples_split': 'Min Samples Split',
            'min_samples_leaf': 'Min Samples Leaf',
            'bootstrap': 'Bootstrap Sampling',
            'max_leaf_nodes': 'Max Leaf Nodes',
            'max_samples': 'Max Samples Ratio',
            'C': 'Regularization Parameter (C)',
            'epsilon': 'Epsilon (SVR)',
            'kernel': 'Kernel Type',
            'degree': 'Polynomial Degree',
            'coef0': 'Independent Term (coef0)',
            'tol': 'Tolerance',
            'hidden_layer_sizes': 'Hidden Layer Sizes',
            'activation': 'Activation Function',
            'solver': 'Solver Algorithm',
            'alpha': 'L2 Regularization',
            'learning_rate_init': 'Initial Learning Rate',
            'max_iter': 'Max Iterations',
            'early_stopping': 'Early Stopping',
            'shrinking': 'Use Shrinking Heuristic'
        }
        return expansions.get(param_name, param_name)

    def run(self):
        """Main execution function"""
        print("="*60)
        print("OPTIMIZATION HISTORY TRACKER")
        print("="*60)

        # Select log file
        log_file = self.select_log_file()
        if not log_file:
            return

        # Parse log file
        print(f"\nParsing log file: {log_file}")
        results = self.parse_log_file(log_file)

        if not results:
            print("No optimization results found in log file.")
            return

        print(f"Found {len(results)} optimization results")

        # Load existing history
        print(f"\nLoading existing history from: {self.excel_file}")
        existing_df = self.load_existing_history()

        # Update history
        print(f"\nUpdating history...")
        updated_df, update_details = self.update_history(existing_df, results)

        # Save updated history
        self.save_history(updated_df, update_details)

def main():
    """Main function to run the optimization history tracker.

    Parameters
    ----------
    None

    Returns
    -------
    None
    """
    tracker = OptimizationHistoryTracker()
    tracker.run()

if __name__ == "__main__":
    main()
