#!/usr/bin/env python3
"""
Optimization History Tracker

This script parses ML training log files to extract optimization results
and maintains a running history of the best parameters for each model/encoding/optimization
combination in an Excel file.
"""

import os
import re
import pandas as pd
import ast
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import numpy as np
import openpyxl

class OptimizationHistoryTracker:
    def __init__(self):
        self.log_dir = "ML/outputs/logs"
        self.excel_file = "ML/hyperparameter_tuning/optimisation_history.xlsx"

                # Define grouped parameters by model type (excluding non-essential params)
        self.xgboost_params = [
            'n_estimators', 'max_depth', 'learning_rate', 'subsample',
            'colsample_bytree', 'colsample_bylevel', 'reg_alpha', 'reg_lambda',
            'gamma', 'min_child_weight'
        ]

        self.random_forest_params = [
            'n_estimators', 'max_depth', 'max_features', 'min_samples_split',
            'min_samples_leaf', 'bootstrap', 'max_leaf_nodes'
        ]

        self.svm_params = [
            'C', 'epsilon', 'kernel', 'degree', 'coef0'
        ]

        self.neural_network_params = [
            'hidden_layer_sizes', 'activation', 'solver', 'alpha',
            'batch_size', 'max_iter', 'early_stopping', 'validation_fraction'
        ]

        # Combined list for compatibility
        self.all_params = (self.xgboost_params + self.random_forest_params +
                          self.svm_params + self.neural_network_params)

        self.three_stage_params = self.all_params.copy()  # Will be filled with NA mostly

    def select_log_file(self) -> str:
        """Interactive log file selection similar to visualization code"""
        log_path = Path(self.log_dir)

        if not log_path.exists():
            print(f"Error: Log directory {self.log_dir} does not exist.")
            return None

        # Get all log files
        log_files = list(log_path.glob("*.log"))

        if not log_files:
            print(f"No log files found in {self.log_dir}")
            return None

        print(f"\nAvailable log files in {self.log_dir}:")
        print("-" * 50)

        for i, log_file in enumerate(log_files, 1):
            file_size = log_file.stat().st_size / (1024 * 1024)  # MB
            print(f"{i:2d}. {log_file.name} ({file_size:.1f} MB)")

        print(f"{len(log_files) + 1:2d}. Custom path")
        print("-" * 50)

        while True:
            try:
                choice = input(f"Select log file (1-{len(log_files) + 1}): ").strip()

                if choice.isdigit():
                    choice_num = int(choice)
                    if 1 <= choice_num <= len(log_files):
                        selected_file = log_files[choice_num - 1]
                        print(f"Selected: {selected_file}")
                        return str(selected_file)
                    elif choice_num == len(log_files) + 1:
                        custom_path = input("Enter custom log file path: ").strip()
                        if os.path.exists(custom_path):
                            return custom_path
                        else:
                            print("File not found. Please try again.")
                    else:
                        print(f"Please enter a number between 1 and {len(log_files) + 1}")
                else:
                    print("Please enter a valid number")

            except (ValueError, KeyboardInterrupt):
                print("\nOperation cancelled.")
                return None

    def parse_log_file(self, log_file_path: str) -> List[Dict]:
        """Parse log file to extract optimization results"""
        results = []

        try:
            with open(log_file_path, 'r', encoding='utf-8') as f:
                content = f.read()
        except Exception as e:
            print(f"Error reading log file: {e}")
            return results

        # Split content into sections for each job
        job_sections = self._split_into_job_sections(content)
        print(f"Found {len(job_sections)} job sections in log file")

        for i, section in enumerate(job_sections, 1):
            result = self._parse_job_section(section)
            if result:
                results.append(result)
                print(f"  ✓ Parsed job {i}/{len(job_sections)}: {result['Model']}-{result['Encoding']}-{result['Optimisation']} (MAPE: {result.get('Test MAPE', 'N/A')}%)")
            else:
                print(f"  ✗ Failed to parse job {i}/{len(job_sections)}")

        return results

    def _split_into_job_sections(self, content: str) -> List[str]:
        """Split log content into individual job sections"""
        # Look for job start patterns - handle both with and without comma
        job_pattern = r'Job \d+/\d+: (\w+) for (\w+)[,\s]*\s*Encoding: (\w+) \| Optimization: (\w+)'

        sections = []
        matches = list(re.finditer(job_pattern, content))

        for i, match in enumerate(matches):
            start_pos = match.start()

            # Find the end position (start of next job or end of file)
            if i + 1 < len(matches):
                end_pos = matches[i + 1].start()
            else:
                end_pos = len(content)

            section = content[start_pos:end_pos]
            sections.append(section)

        return sections

    def _parse_job_section(self, section: str) -> Optional[Dict]:
        """Parse a single job section to extract results"""
        result = {}

        # Extract basic info from job header
        job_match = re.search(r'Job \d+/\d+: (\w+) for (\w+)[,\s]*\s*Encoding: (\w+) \| Optimization: (\w+)', section)
        if not job_match:
            return None

        model, target, encoding, optimization = job_match.groups()
        result['Model'] = model
        result['Target'] = target
        result['Encoding'] = encoding
        result['Optimisation'] = optimization

        # Extract Test MAPE
        mape_match = re.search(r'Test MAPE: ([\d.]+)%', section)
        if mape_match:
            result['Test MAPE'] = float(mape_match.group(1))
        else:
            # For keff targets, look for different metrics
            if target == 'keff':
                # For keff, we might not have MAPE, so don't skip
                # but mark MAPE as NaN
                result['Test MAPE'] = np.nan
            else:
                return None  # Skip if no MAPE found for flux

        # Extract additional metrics for all targets
        r2_match = re.search(r'Test R²: ([\d.-]+)', section)
        mae_match = re.search(r'Test MAE: ([\d.e-]+)', section)
        rmse_match = re.search(r'Test RMSE: ([\d.e-]+)', section)
        mse_match = re.search(r'Test MSE: ([\d.e-]+)', section)

        if r2_match:
            result['Test R²'] = float(r2_match.group(1))
        if mae_match:
            result['Test MAE'] = float(mae_match.group(1))
        if rmse_match:
            result['Test RMSE'] = float(rmse_match.group(1))
        if mse_match:
            result['Test MSE'] = float(mse_match.group(1))

        # Extract optimization time
        time_match = re.search(r'Optimization took ([\d.]+) minutes', section)
        if time_match:
            result['Training Time (minutes)'] = float(time_match.group(1))
        else:
            # Look for alternative patterns
            time_match = re.search(r'Total time for \w+ \w+: ([\d.]+) minutes', section)
            if time_match:
                result['Training Time (minutes)'] = float(time_match.group(1))

        # Extract best parameters
        if optimization == 'optuna':
            params = self._extract_optuna_parameters(section)
        elif optimization == 'three_stage':
            params = self._extract_three_stage_parameters(section)
        else:  # none
            params = {}

        result.update(params)

        return result

    def _extract_optuna_parameters(self, section: str) -> Dict:
        """Extract parameters from Optuna optimization section"""
        params = {}

        # Look for best parameters
        params_match = re.search(r'Best parameters found: ({.*?})', section, re.DOTALL)
        if params_match:
            try:
                params_str = params_match.group(1)
                # Clean up the string and parse as Python dict
                params_dict = ast.literal_eval(params_str)

                # Add all found parameters
                for key, value in params_dict.items():
                    params[key] = value

            except Exception as e:
                print(f"Error parsing Optuna parameters: {e}")

        return params

    def _extract_three_stage_parameters(self, section: str) -> Dict:
        """Extract parameters from three-stage optimization section"""
        params = {}

        # Look for final best parameters from Bayesian stage
        # Pattern 1: Similar to Optuna
        params_match = re.search(r'Best parameters found: ({.*?})', section, re.DOTALL)
        if params_match:
            try:
                params_str = params_match.group(1)
                params_dict = ast.literal_eval(params_str)
                for key, value in params_dict.items():
                    params[key] = value
            except Exception as e:
                print(f"Error parsing three-stage parameters: {e}")

        # Pattern 2: Look for "Final best parameters:"
        if not params:
            params_match = re.search(r'Final best parameters: ({.*?})', section, re.DOTALL)
            if params_match:
                try:
                    params_str = params_match.group(1)
                    params_dict = ast.literal_eval(params_str)
                    for key, value in params_dict.items():
                        params[key] = value
                except Exception as e:
                    print(f"Error parsing three-stage final parameters: {e}")

        # If still no parameters found, will be filled with NaN in Excel
        return params

    def load_existing_history(self) -> pd.DataFrame:
        """Load existing optimization history Excel file"""
        if os.path.exists(self.excel_file):
            try:
                df = pd.read_excel(self.excel_file)
                print(f"Loaded existing history with {len(df)} records")
                return df
            except Exception as e:
                print(f"Error loading existing Excel file: {e}")

        # Create empty DataFrame with required columns
        columns = ['Model', 'Target', 'Encoding', 'Optimisation', 'Test MAPE', 'Test R²', 'Test MAE', 'Test RMSE', 'Test MSE', 'Training Time (minutes)']
        columns.extend(self.all_params)
        df = pd.DataFrame(columns=columns)
        print("Created new optimization history")
        return df

    def update_history(self, existing_df: pd.DataFrame, new_results: List[Dict]) -> pd.DataFrame:
        """Update history with new results, only keeping better MAPE scores"""
        updated_df = existing_df.copy()

        updates_made = 0
        additions_made = 0

        for result in new_results:
            # Create identifier for this combination
            identifier_cols = ['Model', 'Target', 'Encoding', 'Optimisation']
            identifier = {col: result[col] for col in identifier_cols}

            # Check if this combination exists
            mask = pd.Series([True] * len(updated_df))
            for col, val in identifier.items():
                mask &= (updated_df[col] == val)

            existing_idx = updated_df[mask].index

            if len(existing_idx) > 0:
                # Combination exists, check if new result is better
                existing_mape = updated_df.loc[existing_idx[0], 'Test MAPE']
                new_mape = result.get('Test MAPE', np.nan)

                is_better = False
                comparison_msg = ""

                # For flux targets, use MAPE (lower is better)
                if identifier['Target'] == 'flux' and not pd.isna(new_mape) and not pd.isna(existing_mape):
                    is_better = new_mape < existing_mape
                    comparison_msg = f"MAPE {existing_mape:.2f}% → {new_mape:.2f}%"
                # For keff targets, use R² (higher is better)
                elif identifier['Target'] == 'keff':
                    existing_r2 = updated_df.loc[existing_idx[0], 'Test R²']
                    new_r2 = result.get('Test R²', np.nan)
                    if not pd.isna(new_r2) and not pd.isna(existing_r2):
                        is_better = new_r2 > existing_r2
                        comparison_msg = f"R² {existing_r2:.4f} → {new_r2:.4f}"
                    elif not pd.isna(new_mape) and not pd.isna(existing_mape):
                        # Fallback to MAPE if R² not available
                        is_better = new_mape < existing_mape
                        comparison_msg = f"MAPE {existing_mape:.2f}% → {new_mape:.2f}%"

                if is_better:
                    # Update with better result
                    for col, val in result.items():
                        if col in updated_df.columns:
                            updated_df.loc[existing_idx[0], col] = val
                    updates_made += 1
                    print(f"Updated {identifier['Model']}-{identifier['Encoding']}-{identifier['Optimisation']}: {comparison_msg}")
                else:
                    print(f"Skipped {identifier['Model']}-{identifier['Encoding']}-{identifier['Optimisation']}: "
                          f"New result not better ({comparison_msg})")
            else:
                # New combination, add it
                new_row_data = {}
                for col in updated_df.columns:
                    new_row_data[col] = result.get(col, np.nan)

                # Create a proper DataFrame row instead of Series to avoid warnings
                new_row_df = pd.DataFrame([new_row_data])
                updated_df = pd.concat([updated_df, new_row_df], ignore_index=True)
                additions_made += 1
                print(f"Added new combination: {identifier['Model']}-{identifier['Encoding']}-{identifier['Optimisation']} "
                      f"(MAPE: {result.get('Test MAPE', 'N/A')}%)")

        print(f"\nSummary: {updates_made} records updated, {additions_made} records added")
        return updated_df

    def save_history(self, df: pd.DataFrame):
        """Save updated history to Excel file with proper formatting and separate sheets"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Ensure directory exists
            os.makedirs(os.path.dirname(self.excel_file), exist_ok=True)

            # Sort by Model, Target, Encoding, Optimisation for better organization
            df_sorted = df.sort_values(['Model', 'Target', 'Encoding', 'Optimisation']).reset_index(drop=True)

            # Split into flux and keff sheets
            flux_df = df_sorted[df_sorted['Target'] == 'flux'].copy()
            keff_df = df_sorted[df_sorted['Target'] == 'keff'].copy()

            # Remove Target column since it's now separated by sheet
            if 'Target' in flux_df.columns:
                flux_df = flux_df.drop('Target', axis=1)
            if 'Target' in keff_df.columns:
                keff_df = keff_df.drop('Target', axis=1)

            # Create Excel writer
            with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                # Write flux sheet
                if not flux_df.empty:
                    flux_df.to_excel(writer, sheet_name='Flux Results', index=False)
                    self._format_sheet(writer.sheets['Flux Results'], flux_df, 'flux')

                # Write keff sheet
                if not keff_df.empty:
                    keff_df.to_excel(writer, sheet_name='Keff Results', index=False)
                    self._format_sheet(writer.sheets['Keff Results'], keff_df, 'keff')

            print(f"\nOptimization history saved to: {self.excel_file}")
            print(f"Flux records: {len(flux_df)}, Keff records: {len(keff_df)}")

        except Exception as e:
            print(f"Error saving Excel file: {e}")

    def _format_sheet(self, ws, df, target_type):
        """Format Excel sheet with colors, grouping, and expandable columns"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Define colors
        colors = {
            'xgboost': PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # Light blue
            'random_forest': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light green
            'svm': PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid"),  # Light orange
            'neural_network': PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid"),  # Light purple
            'param_header': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),  # Light grey
            'param_cell': PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Very light grey
        }

        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Format headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Group columns and apply parameter header formatting
        param_start_col = None
        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)

            # Find where parameters start
            if col_name in self.all_params and param_start_col is None:
                param_start_col = col_idx

            # Format parameter headers
            if col_name in self.all_params:
                # Expand parameter column names
                expanded_name = self._expand_parameter_name(col_name)
                ws[f"{col_letter}1"].value = expanded_name
                ws[f"{col_letter}1"].fill = colors['param_header']
                ws[f"{col_letter}1"].font = Font(bold=True, color="333333")

        # Format data rows
        for row_idx in range(2, len(df) + 2):
            model = df.iloc[row_idx - 2]['Model'].lower()

            # Apply model-specific row coloring
            model_color = colors.get(model, PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))

            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                cell = ws[f"{col_letter}{row_idx}"]

                # Apply model color to main columns
                if col_name not in self.all_params:
                    cell.fill = model_color
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Parameter columns get grey background
                    cell.fill = colors['param_cell']
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Format numeric values
                if col_name in ['Test MAPE', 'Test R²', 'Test MAE', 'Test RMSE', 'Test MSE']:
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        if 'R²' in col_name:
                            cell.number_format = '0.0000'
                        elif 'MAPE' in col_name:
                            cell.number_format = '0.00'
                        else:
                            cell.number_format = '0.00E+00'
                elif col_name == 'Training Time (minutes)':
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0'

        # Create parameter groups with borders
        self._create_parameter_groups(ws, df)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                except:
                    pass

            # Set width with limits
            adjusted_width = min(max(max_length + 2, 10), 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _expand_parameter_name(self, param_name):
        """Expand parameter names to be more descriptive"""
        expansions = {
            'n_estimators': 'Number of Estimators',
            'max_depth': 'Maximum Depth',
            'learning_rate': 'Learning Rate',
            'subsample': 'Subsample Ratio',
            'colsample_bytree': 'Column Sample by Tree',
            'colsample_bylevel': 'Column Sample by Level',
            'reg_alpha': 'L1 Regularization (Alpha)',
            'reg_lambda': 'L2 Regularization (Lambda)',
            'gamma': 'Gamma (Min Split Loss)',
            'min_child_weight': 'Min Child Weight',
            'max_features': 'Max Features',
            'min_samples_split': 'Min Samples Split',
            'min_samples_leaf': 'Min Samples Leaf',
            'bootstrap': 'Bootstrap Sampling',
            'max_leaf_nodes': 'Max Leaf Nodes',
            'C': 'Regularization Parameter (C)',
            'epsilon': 'Epsilon (SVR)',
            'kernel': 'Kernel Type',
            'degree': 'Polynomial Degree',
            'coef0': 'Independent Term (coef0)',
            'hidden_layer_sizes': 'Hidden Layer Sizes',
            'activation': 'Activation Function',
            'solver': 'Solver Algorithm',
            'alpha': 'L2 Regularization',
            'batch_size': 'Batch Size',
            'max_iter': 'Max Iterations',
            'early_stopping': 'Early Stopping',
            'validation_fraction': 'Validation Fraction'
        }
        return expansions.get(param_name, param_name)

    def _create_parameter_groups(self, ws, df):
        """Create visual groups for different parameter types"""
        from openpyxl.styles import Border, Side

        # Define parameter groups
        groups = {
            'XGBoost': self.xgboost_params,
            'Random Forest': self.random_forest_params,
            'SVM': self.svm_params,
            'Neural Network': self.neural_network_params
        }

        # Find column positions for each group
        for group_name, group_params in groups.items():
            group_cols = []
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name in group_params:
                    group_cols.append(col_idx)

            if group_cols:
                # Add subtle borders around parameter groups
                start_col = min(group_cols)
                end_col = max(group_cols)

                border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )

                # Apply border to the group
                for row in range(1, len(df) + 2):
                    for col in range(start_col, end_col + 1):
                        col_letter = get_column_letter(col)
                        ws[f"{col_letter}{row}"].border = border

    def run(self):
        """Main execution function"""
        print("="*60)
        print("OPTIMIZATION HISTORY TRACKER")
        print("="*60)

        # Select log file
        log_file = self.select_log_file()
        if not log_file:
            return

        # Parse log file
        print(f"\nParsing log file: {log_file}")
        results = self.parse_log_file(log_file)

        if not results:
            print("No optimization results found in log file.")
            return

        print(f"Found {len(results)} optimization results")

        # Load existing history
        existing_df = self.load_existing_history()

        # Update history
        updated_df = self.update_history(existing_df, results)

        # Save updated history
        self.save_history(updated_df)

def main():
    tracker = OptimizationHistoryTracker()
    tracker.run()

if __name__ == "__main__":
    main()
