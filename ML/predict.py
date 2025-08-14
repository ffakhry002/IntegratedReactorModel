import sys
import os
import numpy as np
import pandas as pd
from datetime import datetime
import multiprocessing
from joblib import Parallel, delayed
from tqdm import tqdm
import time

# Suppress joblib parallel output messages
os.environ['JOBLIB_VERBOSITY'] = '0'

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from test_execution.model_tester import ReactorModelTester
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import glob
import joblib


class PredictionRunner:
    """Class to handle ML model predictions with user interaction"""

    def __init__(self):
        self.models_dir = None
        self.excel_reports_dir = None
        self.tester = None
        self.available_models = []
        self.model_map = {}
        self.loaded_models_cache = {}  # Cache for loaded models
        self.current_model_key = None  # Track current model to avoid reloading

    def setup_directories(self):
        """Setup output directories"""
        ml_dir = os.path.dirname(os.path.abspath(__file__))
        outputs_dir = os.path.join(ml_dir, "outputs")
        self.models_dir = os.path.join(outputs_dir, "models")
        self.excel_reports_dir = os.path.join(outputs_dir, "excel_reports")

        # Ensure directories exist
        os.makedirs(self.excel_reports_dir, exist_ok=True)

        # Initialize tester
        self.tester = ReactorModelTester(models_dir=self.models_dir)

    def find_and_display_models(self):
        """Find available models and display them to user"""
        print("\nSearching for trained models...")
        self.available_models = self.tester.find_available_models()

        if not self.available_models:
            print(f"No trained models found in '{self.models_dir}' directory.")
            print("Please train models first using main.py")
            return False

        # Group models by type
        flux_models = [m for m in self.available_models if m['model_type'] == 'flux']
        keff_models = [m for m in self.available_models if m['model_type'] == 'keff']

        print(f"\nFound {len(self.available_models)} trained models:")

        # Show file sizes to help identify large models
        total_size = 0
        for model in self.available_models:
            file_size = os.path.getsize(model['filepath']) / (1024 * 1024)  # MB
            total_size += file_size
            if file_size > 50:  # Warn about large models
                print(f"  ⚠️  Large model file: {os.path.basename(model['filepath'])} ({file_size:.1f} MB)")

        if total_size > 100:
            print(f"  Total model size: {total_size:.1f} MB")
            print("  Note: Large models may take longer to load\n")

        model_index = 1
        self.model_map = {}

        # Display flux models by flux mode
        if flux_models:
            print("\nFLUX MODELS:")
            total_flux = [m for m in flux_models if m.get('flux_mode', 'total') == 'total']
            energy_flux = [m for m in flux_models if m.get('flux_mode') == 'energy']
            bin_flux = [m for m in flux_models if m.get('flux_mode') == 'bin']
            thermal_only = [m for m in flux_models if m.get('flux_mode') == 'thermal_only']
            epithermal_only = [m for m in flux_models if m.get('flux_mode') == 'epithermal_only']
            fast_only = [m for m in flux_models if m.get('flux_mode') == 'fast_only']

            for category, desc in [
                 (total_flux, "Total Flux Models"),
                 (energy_flux, "Energy Flux Models"),
                 (bin_flux, "Bin Flux Models"),
                 (thermal_only, "Thermal Only Models"),
                 (epithermal_only, "Epithermal Only Models"),
                 (fast_only, "Fast Only Models")
            ]:
                if category:
                    print(f"  {desc}:")
                    for model in category:
                        print(f"    {model_index}. {model['model_class']} - {model.get('flux_mode', 'total')} flux "
                              f"({model['encoding']}, {model['optimization_method']})")
                        self.model_map[model_index] = model
                        model_index += 1

        if keff_models:
            print("\nK-EFF MODELS:")
            for model in keff_models:
                print(f"    {model_index}. {model['model_class']} - k-eff "
                      f"({model['encoding']}, {model['optimization_method']})")
                self.model_map[model_index] = model
                model_index += 1

        return True

    def get_user_prediction_choices(self):
        """Get user choices for what to predict"""
        choices = {}

        # K-eff prediction
        keff_choice = input("\nDo you want to do k-eff prediction? (y/n): ").strip().lower()
        choices['predict_keff'] = keff_choice == 'y'

        if choices['predict_keff']:
            keff_models = [i for i, m in self.model_map.items() if m['model_type'] == 'keff']
            if not keff_models:
                print("No k-eff models available!")
                choices['predict_keff'] = False
            else:
                print("\nAvailable k-eff models:")
                for i in keff_models:
                    model = self.model_map[i]
                    print(f"    {i}. {model['model_class']} - k-eff "
                        f"({model['encoding']}, {model['optimization_method']})")

                try:
                    keff_model_choice = int(input("Select k-eff model number: ").strip())
                    if keff_model_choice in keff_models:
                        choices['keff_model'] = self.model_map[keff_model_choice]
                    else:
                        print("Invalid model selection!")
                        choices['predict_keff'] = False
                except:
                    print("Invalid input!")
                    choices['predict_keff'] = False

        # Flux predictions
        thermal_choice = input("\nDo you want to predict thermal flux? (y/n): ").strip().lower()
        choices['predict_thermal'] = thermal_choice == 'y'

        epithermal_choice = input("Do you want to predict epithermal flux? (y/n): ").strip().lower()
        choices['predict_epithermal'] = epithermal_choice == 'y'

        fast_choice = input("Do you want to predict fast flux? (y/n): ").strip().lower()
        choices['predict_fast'] = fast_choice == 'y'

        # If all three flux types selected, ask about combining
        if choices['predict_thermal'] and choices['predict_epithermal'] and choices['predict_fast']:
            combine_choice = input("\nDo you want to add the fluxes (thermal + epithermal + fast) or do separate predictions? (add/separate): ").strip().lower()
            choices['combine_fluxes'] = combine_choice == 'add'

            # If adding fluxes, we DON'T need a separate total flux prediction
            if choices['combine_fluxes']:
                choices['predict_total'] = False
                print("Total flux will be calculated as the sum of thermal + epithermal + fast")
            else:
                # Only ask about total flux if doing separate predictions
                total_choice = input("\nDo you want to predict total flux independently? (y/n): ").strip().lower()
                choices['predict_total'] = total_choice == 'y'
        else:
            # If not all three are selected, can't combine them
            choices['combine_fluxes'] = False
            # Ask about total flux prediction
            total_choice = input("\nDo you want to predict total flux? (y/n): ").strip().lower()
            choices['predict_total'] = total_choice == 'y'

        # Get model selections for each flux type
        flux_models = [i for i, m in self.model_map.items() if m['model_type'] == 'flux']

        if choices['predict_thermal']:
            choices['thermal_model'] = self._select_flux_model("thermal", flux_models)
            if not choices['thermal_model']:
                choices['predict_thermal'] = False

        if choices['predict_epithermal']:
            choices['epithermal_model'] = self._select_flux_model("epithermal", flux_models)
            if not choices['epithermal_model']:
                choices['predict_epithermal'] = False

        if choices['predict_fast']:
            choices['fast_model'] = self._select_flux_model("fast", flux_models)
            if not choices['fast_model']:
                choices['predict_fast'] = False

        if choices.get('predict_total', False):
            choices['total_model'] = self._select_flux_model("total", flux_models)
            if not choices['total_model']:
                choices['predict_total'] = False

        return choices

    def _select_flux_model(self, flux_type, flux_models):
        """Helper to select a flux model"""
        print(f"\nAvailable flux models for {flux_type}:")
        relevant_models = []

        for i in flux_models:
            model = self.model_map[i]
            flux_mode = model.get('flux_mode', 'total')

            # Show relevant models based on flux type
            if flux_type == 'thermal' and flux_mode in ['thermal_only', 'energy', 'bin']:
                relevant_models.append(i)
            elif flux_type == 'epithermal' and flux_mode in ['epithermal_only', 'energy', 'bin']:
                relevant_models.append(i)
            elif flux_type == 'fast' and flux_mode in ['fast_only', 'energy', 'bin']:
                relevant_models.append(i)
            elif flux_type == 'total' and flux_mode == 'total':
                relevant_models.append(i)

        if not relevant_models:
            print(f"No suitable models found for {flux_type} flux!")
            return None

        for i in relevant_models:
            model = self.model_map[i]
            print(f"    {i}. {model['model_class']} - {model.get('flux_mode', 'total')} flux "
                  f"({model['encoding']}, {model['optimization_method']})")

        try:
            choice = int(input(f"Select {flux_type} flux model number: ").strip())
            if choice in relevant_models:
                return self.model_map[choice]
            else:
                print("Invalid model selection!")
                return None
        except:
            print("Invalid input!")
            return None

    def get_input_data(self):
        """Get input data for prediction"""
        print("\n" + "-"*40)

        # Look for files in ML/data directory
        data_dir = "ML/data"
        if not os.path.exists(data_dir):
            print(f"Error: Data directory '{data_dir}' not found.")
            return None

        # Get all .txt files in the data directory
        txt_files = [f for f in os.listdir(data_dir) if f.endswith('.txt')]

        if not txt_files:
            print(f"No .txt files found in '{data_dir}' directory.")
            return None

        # Display available files
        print("Available data files:")
        file_map = {}
        for i, filename in enumerate(sorted(txt_files), 1):
            file_path = os.path.join(data_dir, filename)
            file_size = os.path.getsize(file_path)
            print(f"    {i}. {filename} ({file_size} bytes)")
            file_map[i] = os.path.join(data_dir, filename)

        # Get user selection
        try:
            choice = input(f"\nSelect file number (1-{len(txt_files)}): ").strip()
            file_num = int(choice)

            if file_num in file_map:
                selected_file = file_map[file_num]
                print(f"Selected: {selected_file}")
                return selected_file
            else:
                print("Invalid file selection!")
                return None

        except ValueError:
            print("Invalid input! Please enter a number.")
            return None

    def _load_model_cached(self, model_info):
        """Load a model with caching to avoid repeated file I/O"""
        filepath = model_info['filepath']

        # Create a unique key for this model
        model_key = f"{filepath}_{model_info.get('model_type', 'unknown')}"

        # Check if this is the same model we just used
        if self.current_model_key == model_key and filepath in self.loaded_models_cache:
            return self.loaded_models_cache[filepath]

        # Check if already cached but different from current
        if filepath in self.loaded_models_cache:
            self.current_model_key = model_key
            return self.loaded_models_cache[filepath]

        # Clear cache if switching models (memory management)
        if self.current_model_key != model_key:
            self._clear_model_cache()
            self.current_model_key = model_key

        # Load the model silently
        model_data = joblib.load(filepath)
        model_class_name = model_info.get('model_class', 'unknown')

        # Load using appropriate model class
        model_wrapper = None
        if model_class_name == 'xgboost':
            from ML_models import XGBoostReactorModel
            model_wrapper, metadata = XGBoostReactorModel.load_model(filepath)
        elif model_class_name == 'random_forest':
            from ML_models import RandomForestReactorModel
            model_wrapper, metadata = RandomForestReactorModel.load_model(filepath)
        elif model_class_name == 'svm':
            from ML_models import SVMReactorModel
            model_wrapper, metadata = SVMReactorModel.load_model(filepath)
        elif model_class_name == 'neural_net':
            from ML_models import NeuralNetReactorModel
            model_wrapper, metadata = NeuralNetReactorModel.load_model(filepath)

        # Cache the loaded model
        self.loaded_models_cache[filepath] = {
            'model_data': model_data,
            'model_wrapper': model_wrapper,
            'metadata': metadata
        }

        return self.loaded_models_cache[filepath]

    def _clear_model_cache(self):
        """Clear the model cache to free memory"""
        self.loaded_models_cache.clear()

    def _process_single_prediction(self, config_data):
        """Process a single prediction for a specific model and configuration"""
        i, lattice, description, model_info, prediction_type, flux_type = config_data

        try:
            if prediction_type == 'keff':
                return i, self._predict_single_value(lattice, model_info, 'keff'), 'keff', None
            else:  # flux prediction
                flux_values = self._predict_flux_values(lattice, model_info, flux_type)
                return i, flux_values, flux_type, None
        except Exception as e:
            return i, None, prediction_type, str(e)

    def run_predictions(self, choices, test_file):
        """Run predictions grouped by model type for efficiency"""
        print("\nLoading test data...")

        # Parse test data - handle both formats
        lattices = []
        descriptions = []

        with open(test_file, 'r') as f:
            content = f.read()

        # Simple parsing for lattice-only format
        import re

        # Split by RUN entries - more robust pattern
        run_pattern = r'RUN\s+\d+\s*:'
        runs = re.split(run_pattern, content)

        print(f"Found {len(runs)-1} RUN entries in file")

        for i, run in enumerate(runs[1:], 1):  # Skip empty first split
            if not run.strip():
                continue

            # Extract description - more flexible pattern
            desc_match = re.search(r'Description:\s*([^\n]+)', run)
            if desc_match:
                description = desc_match.group(1).strip()
            else:
                print(f"Warning: No description found for RUN {i}")
                description = f"Config_{i}"

            # Extract lattice - more flexible pattern
            lattice_match = re.search(r'core_lattice:\s*(\[.*?\])\s*(?:Success|Modified|$|=)', run, re.DOTALL)

            if lattice_match:
                lattice_str = lattice_match.group(1)
                try:
                    # Clean up the string more thoroughly
                    lattice_str = lattice_str.replace('\n', ' ')
                    lattice_str = re.sub(r'\s+', ' ', lattice_str)
                    lattice_str = lattice_str.strip()

                    # Try to evaluate the lattice
                    lattice_list = eval(lattice_str)

                    # Verify it's a valid 8x8 lattice
                    if len(lattice_list) == 8 and all(len(row) == 8 for row in lattice_list):
                        lattice_array = np.array(lattice_list, dtype='<U10')
                        lattices.append(lattice_array)
                        descriptions.append(description)
                    else:
                        print(f"Invalid lattice dimensions for {description}: {len(lattice_list)}x{len(lattice_list[0]) if lattice_list else 0}")
                except Exception as e:
                    # More detailed error reporting
                    print(f"Failed to parse lattice for {description}")
                    print(f"  Error: {str(e)}")
                    if len(lattice_str) > 100:
                        print(f"  Lattice string (first 100 chars): {lattice_str[:100]}...")
                    else:
                        print(f"  Lattice string: {lattice_str}")
                    continue
            else:
                print(f"No lattice found for {description}")

        print(f"\nSuccessfully loaded {len(lattices)} configurations")

        if len(lattices) == 0:
            print("\nNo valid configurations found!")
            print("Check if your file format matches the expected pattern:")
            print("  RUN N:")
            print("  ----------------------------------------")
            print("  Description: config_name")
            print("  core_lattice: [[...]]")
            return []

        # Initialize results dictionary
        results = []
        for i, lattice in enumerate(lattices):
            description = descriptions[i] if i < len(descriptions) else f"Config_{i+1}"
            results.append({
                'config_id': f"Config_{i+1}",
                'description': description
            })

        # Rest of the function remains the same...
        # Process predictions grouped by model type
        if choices.get('predict_keff') and choices.get('keff_model'):
            model_info = choices['keff_model']
            print(f"\nProcessing k-eff predictions with {model_info['model_class']}...")

            print(f"Loading k-eff model...")
            self._load_model_cached(model_info)

            config_data = [(i, lattices[i], descriptions[i], model_info, 'keff', None)
                        for i in range(len(lattices))]

            keff_results = []
            for data in tqdm(config_data, desc=f"K-eff ({model_info['model_class']})"):
                result = self._process_single_prediction(data)
                keff_results.append(result)

            for idx, value, pred_type, error in keff_results:
                if error:
                    results[idx]['keff'] = 'Error'
                else:
                    results[idx]['keff'] = value

        # Process each flux type
        for flux_type in ['thermal', 'epithermal', 'fast', 'total']:
            if choices.get(f'predict_{flux_type}') and choices.get(f'{flux_type}_model'):
                model_info = choices[f'{flux_type}_model']
                print(f"\nProcessing {flux_type} flux predictions with {model_info['model_class']}...")

                print(f"Loading {flux_type} flux model...")
                self._load_model_cached(model_info)

                config_data = [(i, lattices[i], descriptions[i], model_info, 'flux', flux_type)
                            for i in range(len(lattices))]

                flux_results = []
                for data in tqdm(config_data, desc=f"{flux_type.capitalize()} flux ({model_info['model_class']})"):
                    result = self._process_single_prediction(data)
                    flux_results.append(result)

                for idx, flux_values, pred_flux_type, error in flux_results:
                    if not error and flux_values:
                        for label, value in flux_values.items():
                            if label.startswith('I_'):
                                label_num = int(label.split('_')[1])
                                results[idx][f'I_{label_num}_{flux_type}'] = value

        self._clear_model_cache()

        print("\nCalculating aggregated values...")
        for i, result in enumerate(results):
            self._calculate_aggregated_values(result, choices, lattices[i])

        return results

    def _calculate_aggregated_values(self, result, choices, lattice):
        """Calculate percentages and averages for a single result"""
        # Find all irradiation labels in lattice
        irr_labels = []
        for i in range(lattice.shape[0]):
            for j in range(lattice.shape[1]):
                if lattice[i, j].startswith('I_'):
                    if lattice[i, j] not in irr_labels:
                        irr_labels.append(lattice[i, j])
        irr_labels.sort()

        # Calculate percentages if all three energy groups are present
        if choices.get('predict_thermal') and choices.get('predict_epithermal') and choices.get('predict_fast'):
            for label in irr_labels:
                label_num = int(label.split('_')[1])

                thermal_key = f'I_{label_num}_thermal'
                epithermal_key = f'I_{label_num}_epithermal'
                fast_key = f'I_{label_num}_fast'

                if all(key in result for key in [thermal_key, epithermal_key, fast_key]):
                    thermal_val = result[thermal_key]
                    epithermal_val = result[epithermal_key]
                    fast_val = result[fast_key]

                    total_flux = thermal_val + epithermal_val + fast_val

                    if total_flux > 0:
                        result[f'I_{label_num}_thermal_percent'] = (thermal_val / total_flux) * 100
                        result[f'I_{label_num}_epithermal_percent'] = (epithermal_val / total_flux) * 100
                        result[f'I_{label_num}_fast_percent'] = (fast_val / total_flux) * 100
                        result[f'I_{label_num}_total_flux'] = total_flux

        # Calculate averages
        for flux_type in ['thermal', 'epithermal', 'fast', 'total']:
            if choices.get(f'predict_{flux_type}', False):
                values = []
                for label in irr_labels:
                    label_num = int(label.split('_')[1])
                    key = f'I_{label_num}_{flux_type}'
                    if key in result:
                        values.append(result[key])

                if values:
                    result[f'average_{flux_type}_flux'] = np.mean(values)

        # Calculate total flux statistics if we have combined flux
        if 'I_1_total_flux' in result:  # Check if we calculated total flux
            total_values = []
            for label in irr_labels:
                label_num = int(label.split('_')[1])
                key = f'I_{label_num}_total_flux'
                if key in result:
                    total_values.append(result[key])

            if total_values:
                result['average_total_flux'] = np.mean(total_values)
                result['min_total_flux'] = np.min(total_values)
                result['max_total_flux'] = np.max(total_values)

    def _predict_single_value(self, lattice, model_info, value_type):
        """Predict a single value (like k-eff)"""
        # Encode lattice
        encode_result = self.tester.encode_lattice(lattice, model_info['encoding'])

        if len(encode_result) == 3:
            features, irr_positions, position_order = encode_result
        else:
            features, irr_positions = encode_result
            position_order = sorted(irr_positions)

        features = features.reshape(1, -1)

        # Load model using cache
        cached_model = self._load_model_cached(model_info)
        model_data = cached_model['model_data']
        model_wrapper = cached_model['model_wrapper']

        if model_wrapper:
            if value_type == 'keff':
                prediction = model_wrapper.predict_keff(features)
            else:
                prediction = model_wrapper.predict_flux(features)
        else:
            # Fallback to raw model
            model = model_data.get('model')
            if 'scaler' in model_data:
                features = model_data['scaler'].transform(features)
            prediction = model.predict(features)

        return prediction[0] if hasattr(prediction, '__len__') else prediction

    def _predict_flux_values(self, lattice, model_info, flux_type):
        """Predict flux values for all irradiation positions"""
        # Encode lattice
        encode_result = self.tester.encode_lattice(lattice, model_info['encoding'])

        if len(encode_result) == 3:
            features, irr_positions, position_order = encode_result
        else:
            features, irr_positions = encode_result
            position_order = sorted(irr_positions)

        features = features.reshape(1, -1)

        # Load model using cache
        cached_model = self._load_model_cached(model_info)
        model_data = cached_model['model_data']
        model_wrapper = cached_model['model_wrapper']

        if model_wrapper:
            flux_pred = model_wrapper.predict_flux(features)
        else:
            # Fallback to raw model
            model = model_data.get('model')
            if 'scaler' in model_data:
                features = model_data['scaler'].transform(features)
            flux_pred = model.predict(features)

        if len(flux_pred.shape) == 1:
            flux_pred = flux_pred.reshape(1, -1)
        flux_pred = flux_pred[0]  # Get first (and only) sample

        # Process based on flux mode
        flux_mode = model_info.get('flux_mode', 'total')

        if flux_mode == 'total' or flux_mode.endswith('_only'):
            # Handle scaling for these modes
            use_log_flux = model_info.get('use_log_flux', model_data.get('use_log_flux', False))
            flux_scale = model_info.get('flux_scale', model_data.get('flux_scale', 1e14))

            if use_log_flux:
                flux_pred_original = 10 ** flux_pred
            else:
                flux_pred_original = flux_pred * flux_scale

            # Map predictions to positions
            position_to_flux_pred = {}
            for idx, pos in enumerate(position_order):
                if idx < len(flux_pred_original):
                    position_to_flux_pred[pos] = flux_pred_original[idx]

            # Map positions to labels
            label_to_flux_pred = {}
            for i in range(lattice.shape[0]):
                for j in range(lattice.shape[1]):
                    if lattice[i, j].startswith('I_') and (i, j) in position_to_flux_pred:
                        label = lattice[i, j]
                        label_to_flux_pred[label] = position_to_flux_pred[(i, j)]

            return label_to_flux_pred

        elif flux_mode in ['energy', 'bin']:
            # Handle energy/bin modes with multiple energy groups
            # TODO: Implement proper energy group handling
            return {'raw_predictions': flux_pred}

        return {}

    def create_excel_report(self, results, choices):
        """Create Excel report with dynamic columns"""
        print("\n" + "-"*40)
        output_file = input("Output Excel filename (press Enter for timestamp): ").strip()
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"predictions_{timestamp}.xlsx"
        elif not output_file.endswith('.xlsx'):
            output_file += '.xlsx'

        output_path = os.path.join(self.excel_reports_dir, output_file)

        # Build column order dynamically
        columns = ['config_id', 'description']

        # Add k-eff if predicted
        if choices['predict_keff']:
            columns.append('keff')

        # Find max irradiation position number
        max_irr_pos = 0
        for result in results:
            for key in result.keys():
                if key.startswith('I_') and '_' in key:
                    try:
                        pos_num = int(key.split('_')[1])
                        max_irr_pos = max(max_irr_pos, pos_num)
                    except:
                        pass

        # Add flux columns in the specified order
        for flux_type in ['thermal', 'epithermal', 'fast']:
            if choices.get(f'predict_{flux_type}', False):
                for i in range(1, max_irr_pos + 1):
                    col_name = f'I_{i}_{flux_type}'
                    if any(col_name in result for result in results):
                        columns.append(col_name)

        # Add total flux columns if calculated
        for i in range(1, max_irr_pos + 1):
            col_name = f'I_{i}_total_flux'
            if any(col_name in result for result in results):
                columns.append(col_name)

        # Add percentage columns if all three energy groups
        if choices.get('predict_thermal') and choices.get('predict_epithermal') and choices.get('predict_fast'):
            for i in range(1, max_irr_pos + 1):
                for flux_type in ['thermal', 'epithermal', 'fast']:
                    col_name = f'I_{i}_{flux_type}_percent'
                    if any(col_name in result for result in results):
                        columns.append(col_name)

        # Add total flux if predicted separately
        if choices.get('predict_total', False):
            for i in range(1, max_irr_pos + 1):
                col_name = f'I_{i}_total'
                if any(col_name in result for result in results):
                    columns.append(col_name)

        # Add average columns
        for flux_type in ['thermal', 'epithermal', 'fast', 'total']:
            avg_col = f'average_{flux_type}_flux'
            if any(avg_col in result for result in results):
                columns.append(avg_col)

        # Add min/max total flux
        if any('min_total_flux' in result for result in results):
            columns.extend(['min_total_flux', 'max_total_flux'])

        # Filter results to only include existing columns
        filtered_results = []
        for result in results:
            filtered_result = {}
            for col in columns:
                filtered_result[col] = result.get(col, '')
            filtered_results.append(filtered_result)

        # Create DataFrame and save to Excel
        df = pd.DataFrame(filtered_results)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Predictions', index=False)

            # Format the sheet
            ws = writer.sheets['Predictions']
            self._format_excel_sheet(ws, df)

        print(f"\nExcel report saved to: {output_file}")
        return output_path

    def _format_excel_sheet(self, ws, df):
        """Apply formatting to Excel sheet"""
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format data columns
        for col_idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)

            # Format scientific notation for flux columns
            if ('flux' in col or col.startswith('I_')) and 'percent' not in col and col != 'config_id':
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00E+00'

            # Format percentage columns
            elif 'percent' in col:
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'

            # Format k-eff values
            elif col == 'keff':
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.000000'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width


def main():
    """Main function to run prediction pipeline"""
    print("\n" + "="*60)
    print("NUCLEAR REACTOR MODEL PREDICTION SYSTEM")
    print("="*60)

    predictor = PredictionRunner()

    try:
        # Setup directories
        predictor.setup_directories()

        # Find and display models
        if not predictor.find_and_display_models():
            return

        # Get user choices
        choices = predictor.get_user_prediction_choices()

        # Check if any predictions were selected
        if not any([choices.get('predict_keff', False),
                   choices.get('predict_thermal', False),
                   choices.get('predict_epithermal', False),
                   choices.get('predict_fast', False),
                   choices.get('predict_total', False)]):
            print("No predictions selected. Exiting.")
            return

        # Get input data
        test_file = predictor.get_input_data()
        if not test_file:
            return

        # Run predictions
        print("\nRunning predictions...")
        results = predictor.run_predictions(choices, test_file)

        if not results:
            print("No results generated.")
            return

        # Create Excel report
        output_path = predictor.create_excel_report(results, choices)

        print("\n" + "="*60)
        print("PREDICTION COMPLETE")
        print("="*60)
        print(f"Results saved to: {os.path.basename(output_path)}")

    except KeyboardInterrupt:
        print("\n\nPrediction cancelled by user.")
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
