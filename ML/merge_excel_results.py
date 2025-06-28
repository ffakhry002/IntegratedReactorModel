#!/usr/bin/env python3
"""
Excel Results Merger Utility
Merges different types of model results into combined Excel files
Supports:
1. All energy merge (thermal + epithermal + fast)
2. Total + K-eff merge
3. K-eff + Multi-energy merge
4. K-eff + Energy merge (thermal + epithermal + fast + keff)
"""

import pandas as pd
import os
import sys
from datetime import datetime
from pathlib import Path

def select_excel_file(prompt_message):
    """Interactive Excel file selection from ML/outputs/excel_reports"""
    excel_dir = "ML/outputs/excel_reports"
    excel_path = Path(excel_dir)

    if not excel_path.exists():
        print(f"Error: Excel directory {excel_dir} does not exist.")
        return None

    # Get all Excel files
    excel_files = list(excel_path.glob("*.xlsx"))

    if not excel_files:
        print(f"No Excel files found in {excel_dir}")
        return None

    print(f"\n{prompt_message}")
    print(f"Available Excel files in {excel_dir}:")
    print("-" * 60)

    for i, excel_file in enumerate(excel_files, 1):
        file_size = excel_file.stat().st_size / (1024 * 1024)  # MB
        mod_time = datetime.fromtimestamp(excel_file.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
        print(f"{i:2d}. {excel_file.name}")
        print(f"     Size: {file_size:.1f} MB | Modified: {mod_time}")

    print(f"{len(excel_files) + 1:2d}. Custom path")
    print("-" * 60)

    while True:
        try:
            choice = input(f"Select Excel file (1-{len(excel_files) + 1}): ").strip()

            if choice.isdigit():
                choice_num = int(choice)
                if 1 <= choice_num <= len(excel_files):
                    selected_file = excel_files[choice_num - 1]
                    print(f"Selected: {selected_file.name}")
                    return str(selected_file)
                elif choice_num == len(excel_files) + 1:
                    custom_path = input("Enter custom Excel file path: ").strip()
                    if os.path.exists(custom_path):
                        return custom_path
                    else:
                        print("File not found. Please try again.")
                else:
                    print(f"Please enter a number between 1 and {len(excel_files) + 1}")
            else:
                print("Please enter a valid number")

        except (ValueError, KeyboardInterrupt):
            print("\nOperation cancelled.")
            return None

def load_excel_data(filepath):
    """Load Excel file and return dataframe"""
    try:
        df = pd.read_excel(filepath, sheet_name='Test Results')
        print(f"✓ Loaded {len(df)} rows from {os.path.basename(filepath)}")
        return df
    except Exception as e:
        print(f"✗ Error loading {filepath}: {e}")
        return None

def detect_data_type(df):
    """Detect what type of data is in the dataframe"""
    data_types = []

    # Check for flux modes
    if 'flux_mode' in df.columns:
        flux_modes = df['flux_mode'].unique()
        data_types.extend(flux_modes)
    else:
        # Infer from columns
        cols = df.columns
        if any('_thermal_' in col for col in cols):
            data_types.append('thermal_only')
        if any('_epithermal_' in col for col in cols):
            data_types.append('epithermal_only')
        if any('_fast_' in col for col in cols):
            data_types.append('fast_only')
        if any(col.startswith('I_') and '_predicted' in col and
               not any(energy in col for energy in ['_thermal_', '_epithermal_', '_fast_'])
               for col in cols):
            data_types.append('total')

    # Check for k-eff
    if any('keff' in col.lower() for col in df.columns):
        data_types.append('keff')

    return data_types

def create_merge_key(row):
    """Create a unique key for matching configurations across files"""
    # Use config_id, model_class, encoding, and optimization_method
    key_parts = [
        str(row.get('config_id', '')),
        str(row.get('model_class', '')),
        str(row.get('encoding', '')),
        str(row.get('optimization_method', ''))
    ]
    return '|'.join(key_parts)

def save_merged_excel(df, output_path):
    """Save merged dataframe to Excel with proper formatting for mixed data"""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    print(f"💾 Saving merged results to Excel...")

    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write the data
        df.to_excel(writer, sheet_name='Merged Results', index=False)

        # Get the worksheet for formatting
        ws = writer.sheets['Merged Results']

        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format columns based on content
        for col_idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)

            # Format in_training column if present
            if col == 'in_training':
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.value == 'T':
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                    else:
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red

            # Format percentage/error columns
            elif 'rel_error' in col or 'Error' in col or 'mape' in col.lower():
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'

            # Format scientific notation for flux columns
            elif (('flux' in col or col.startswith('I_')) and
                  ('_real' in col or '_predicted' in col) and
                  'error' not in col):
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00E+00'

            # Format k-eff values
            elif 'keff' in col and 'error' not in col:
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.000000'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value != 'N/A' else 'N/A'
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass

            # Limit width for readability
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    print(f"✅ Excel file saved with proper formatting for merged data")

def merge_all_energy(thermal_df, epithermal_df, fast_df):
    """Merge thermal, epithermal, and fast results into multi-energy format"""
    print("\n🔄 Performing all-energy merge...")

    # Add merge keys
    thermal_df['merge_key'] = thermal_df.apply(create_merge_key, axis=1)
    epithermal_df['merge_key'] = epithermal_df.apply(create_merge_key, axis=1)
    fast_df['merge_key'] = fast_df.apply(create_merge_key, axis=1)

    # Get common keys
    common_keys = set(thermal_df['merge_key']) & set(epithermal_df['merge_key']) & set(fast_df['merge_key'])
    print(f"Found {len(common_keys)} common configurations")

    # Start with thermal as base
    merged_df = thermal_df[thermal_df['merge_key'].isin(common_keys)].copy()

    # Rename thermal columns
    thermal_cols = [col for col in merged_df.columns if '_thermal_' in col or
                    (col.startswith('I_') and not any(e in col for e in ['config_id', 'description', 'model_', 'encoding', 'optimization']))]

    for col in thermal_cols:
        if '_thermal_' not in col and col.startswith('I_'):
            # Rename I_1_real to I_1_thermal_real, etc.
            new_col = col.replace('I_', 'I_').replace('_real', '_thermal_real').replace('_predicted', '_thermal_predicted').replace('_rel_error', '_thermal_rel_error')
            if 'avg_' in col:
                new_col = col.replace('avg_', 'avg_thermal_')
            if 'mape_' in col:
                new_col = col.replace('mape_', 'mape_thermal_')
            merged_df.rename(columns={col: new_col}, inplace=True)

    # Add epithermal columns
    epithermal_subset = epithermal_df[epithermal_df['merge_key'].isin(common_keys)].set_index('merge_key')
    merged_df = merged_df.set_index('merge_key')

    epithermal_cols = [col for col in epithermal_subset.columns if '_epithermal_' in col or
                      (col.startswith('I_') and not any(e in col for e in ['config_id', 'description', 'model_', 'encoding', 'optimization']))]

    for col in epithermal_cols:
        if '_epithermal_' not in col and col.startswith('I_'):
            # Rename columns
            new_col = col.replace('I_', 'I_').replace('_real', '_epithermal_real').replace('_predicted', '_epithermal_predicted').replace('_rel_error', '_epithermal_rel_error')
            if 'avg_' in col:
                new_col = col.replace('avg_', 'avg_epithermal_')
            if 'mape_' in col:
                new_col = col.replace('mape_', 'mape_epithermal_')
            merged_df[new_col] = epithermal_subset[col]
        else:
            merged_df[col] = epithermal_subset[col]

    # Add fast columns
    fast_subset = fast_df[fast_df['merge_key'].isin(common_keys)].set_index('merge_key')

    fast_cols = [col for col in fast_subset.columns if '_fast_' in col or
                (col.startswith('I_') and not any(e in col for e in ['config_id', 'description', 'model_', 'encoding', 'optimization']))]

    for col in fast_cols:
        if '_fast_' not in col and col.startswith('I_'):
            # Rename columns
            new_col = col.replace('I_', 'I_').replace('_real', '_fast_real').replace('_predicted', '_fast_predicted').replace('_rel_error', '_fast_rel_error')
            if 'avg_' in col:
                new_col = col.replace('avg_', 'avg_fast_')
            if 'mape_' in col:
                new_col = col.replace('mape_', 'mape_fast_')
            merged_df[new_col] = fast_subset[col]
        else:
            merged_df[col] = fast_subset[col]

    # Calculate total columns
    for i in range(1, 5):
        thermal_real = f'I_{i}_thermal_real'
        epithermal_real = f'I_{i}_epithermal_real'
        fast_real = f'I_{i}_fast_real'

        thermal_pred = f'I_{i}_thermal_predicted'
        epithermal_pred = f'I_{i}_epithermal_predicted'
        fast_pred = f'I_{i}_fast_predicted'

        if all(col in merged_df.columns for col in [thermal_real, epithermal_real, fast_real]):
            merged_df[f'I_{i}_total_real'] = merged_df[thermal_real] + merged_df[epithermal_real] + merged_df[fast_real]
            merged_df[f'I_{i}_total_predicted'] = merged_df[thermal_pred] + merged_df[epithermal_pred] + merged_df[fast_pred]

            # Calculate relative error
            total_real = merged_df[f'I_{i}_total_real']
            total_pred = merged_df[f'I_{i}_total_predicted']
            merged_df[f'I_{i}_total_rel_error'] = abs((total_pred - total_real) / total_real * 100).where(total_real != 0, 0)

    # Update flux_mode
    merged_df['flux_mode'] = 'energy'

    # Reset index and clean up
    merged_df = merged_df.reset_index(drop=True)
    if 'merge_key' in merged_df.columns:
        merged_df = merged_df.drop('merge_key', axis=1)

        print(f"✓ Created multi-energy dataset with {len(merged_df)} rows")
    return merged_df

def merge_total_keff(total_df, keff_df):
    """Merge total flux and k-eff results"""
    print("\n🔄 Performing total + k-eff merge...")

    # Add merge keys
    total_df['merge_key'] = total_df.apply(create_merge_key, axis=1)
    keff_df['merge_key'] = keff_df.apply(create_merge_key, axis=1)

    # Get common keys
    common_keys = set(total_df['merge_key']) & set(keff_df['merge_key'])
    print(f"Found {len(common_keys)} common configurations")

    # Start with total as base
    merged_df = total_df[total_df['merge_key'].isin(common_keys)].copy()

    # Add k-eff columns
    keff_subset = keff_df[keff_df['merge_key'].isin(common_keys)].set_index('merge_key')
    merged_df = merged_df.set_index('merge_key')

    keff_cols = ['keff_real', 'keff_predicted', 'keff_rel_error', 'mape_keff']
    for col in keff_cols:
        if col in keff_subset.columns:
            merged_df[col] = keff_subset[col]

    # Reset index and clean up
    merged_df = merged_df.reset_index(drop=True)
    if 'merge_key' in merged_df.columns:
        merged_df = merged_df.drop('merge_key', axis=1)

        print(f"✓ Created total+keff dataset with {len(merged_df)} rows")
    return merged_df

def show_menu():
    """Display the merge options menu"""
    print("\n" + "="*60)
    print("EXCEL RESULTS MERGER")
    print("="*60)
    print("\nMerge Options:")
    print("1. All Energy Merge (thermal + epithermal + fast → multi-energy)")
    print("2. Total + K-eff Merge")
    print("3. K-eff + Multi-energy Merge (existing multi-energy + keff)")
    print("4. K-eff + Energy Merge (thermal + epithermal + fast + keff)")
    print("5. Exit")

    return input("\nSelect merge option (1-5): ").strip()

def get_output_path(default_name):
    """Get output file path from user"""
    output_name = input(f"\nOutput filename (default: {default_name}): ").strip()
    if not output_name:
        output_name = default_name
    if not output_name.endswith('.xlsx'):
        output_name += '.xlsx'

    # Create outputs directory if it doesn't exist
    os.makedirs('ML/outputs/excel_reports', exist_ok=True)

    return os.path.join('ML/outputs/excel_reports', output_name)

def main():
    """Main merge utility"""
    while True:
        choice = show_menu()

        if choice == '1':
            # All Energy Merge
            print("\n📁 Please select the three energy Excel files:")
            thermal_file = select_excel_file("📁 Select thermal flux Excel file:")
            if not thermal_file:
                continue
            epithermal_file = select_excel_file("📁 Select epithermal flux Excel file:")
            if not epithermal_file:
                continue
            fast_file = select_excel_file("📁 Select fast flux Excel file:")
            if not fast_file:
                continue

            # Load data
            thermal_df = load_excel_data(thermal_file)
            epithermal_df = load_excel_data(epithermal_file)
            fast_df = load_excel_data(fast_file)

            if all([thermal_df is not None, epithermal_df is not None, fast_df is not None]):
                # Verify data types
                print("\nVerifying data types...")
                thermal_types = detect_data_type(thermal_df)
                epithermal_types = detect_data_type(epithermal_df)
                fast_types = detect_data_type(fast_df)

                print(f"Thermal file contains: {thermal_types}")
                print(f"Epithermal file contains: {epithermal_types}")
                print(f"Fast file contains: {fast_types}")

                # Perform merge
                merged_df = merge_all_energy(thermal_df, epithermal_df, fast_df)

                # Save
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = get_output_path(f"multi_energy_merged_{timestamp}.xlsx")

                # Save using simple Excel export (custom function for merged data)
                save_merged_excel(merged_df, output_path)

                print(f"\n✅ Saved merged results to: {output_path}")

        elif choice == '2':
            # Total + K-eff Merge
            print("\n📁 Please select the two Excel files:")
            total_file = select_excel_file("📁 Select total flux Excel file:")
            if not total_file:
                continue
            keff_file = select_excel_file("📁 Select K-eff Excel file:")
            if not keff_file:
                continue

            # Load data
            total_df = load_excel_data(total_file)
            keff_df = load_excel_data(keff_file)

            if all([total_df is not None, keff_df is not None]):
                # Verify data types
                print("\nVerifying data types...")
                total_types = detect_data_type(total_df)
                keff_types = detect_data_type(keff_df)

                print(f"Total flux file contains: {total_types}")
                print(f"K-eff file contains: {keff_types}")

                # Perform merge
                merged_df = merge_total_keff(total_df, keff_df)

                # Save
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = get_output_path(f"total_keff_merged_{timestamp}.xlsx")

                save_merged_excel(merged_df, output_path)

                print(f"\n✅ Saved merged results to: {output_path}")

        elif choice == '3':
            # K-eff + Multi-energy Merge
            print("\n📁 Please select the two Excel files:")
            multi_file = select_excel_file("📁 Select multi-energy Excel file:")
            if not multi_file:
                continue
            keff_file = select_excel_file("📁 Select K-eff Excel file:")
            if not keff_file:
                continue

            # Load data
            multi_df = load_excel_data(multi_file)
            keff_df = load_excel_data(keff_file)

            if all([multi_df is not None, keff_df is not None]):
                # Verify data types
                print("\nVerifying data types...")
                multi_types = detect_data_type(multi_df)
                keff_types = detect_data_type(keff_df)

                print(f"Multi-energy file contains: {multi_types}")
                print(f"K-eff file contains: {keff_types}")

                # Perform merge (same as total+keff)
                merged_df = merge_total_keff(multi_df, keff_df)

                # Save
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = get_output_path(f"multi_energy_keff_merged_{timestamp}.xlsx")

                save_merged_excel(merged_df, output_path)

                print(f"\n✅ Saved merged results to: {output_path}")

        elif choice == '4':
            # K-eff + Energy Merge (all four)
            print("\n📁 Please select all four Excel files:")
            thermal_file = select_excel_file("📁 Select thermal flux Excel file:")
            if not thermal_file:
                continue
            epithermal_file = select_excel_file("📁 Select epithermal flux Excel file:")
            if not epithermal_file:
                continue
            fast_file = select_excel_file("📁 Select fast flux Excel file:")
            if not fast_file:
                continue
            keff_file = select_excel_file("📁 Select K-eff Excel file:")
            if not keff_file:
                continue

            # Load data
            thermal_df = load_excel_data(thermal_file)
            epithermal_df = load_excel_data(epithermal_file)
            fast_df = load_excel_data(fast_file)
            keff_df = load_excel_data(keff_file)

            if all([thermal_df is not None, epithermal_df is not None,
                    fast_df is not None, keff_df is not None]):
                # First merge energy
                print("\nStep 1: Merging energy components...")
                energy_merged = merge_all_energy(thermal_df, epithermal_df, fast_df)

                # Then add k-eff
                print("\nStep 2: Adding k-eff data...")
                final_merged = merge_total_keff(energy_merged, keff_df)

                # Save
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = get_output_path(f"full_energy_keff_merged_{timestamp}.xlsx")

                save_merged_excel(final_merged, output_path)

                print(f"\n✅ Saved merged results to: {output_path}")

        elif choice == '5':
            print("\nExiting...")
            break

        else:
            print("\n❌ Invalid choice. Please select 1-5.")

if __name__ == "__main__":
    main()
