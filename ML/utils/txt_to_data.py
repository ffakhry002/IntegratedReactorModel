import numpy as np
import re
from typing import List, Tuple, Dict
import ast
import pandas as pd
from datetime import datetime
import os

def parse_reactor_data(filename: str) -> Tuple[List[np.ndarray], List[Dict], List[float], List[str], List[Dict]]:
    """
    Parse reactor configuration data from text file with validation.
    Returns: (lattices, flux_data, k_effectives, descriptions, energy_groups)
    """
    lattices = []
    flux_data = []
    k_effectives = []
    descriptions = []
    energy_groups = []  # NEW: Store energy group percentages
    all_warnings = []

    with open(filename, 'r') as f:
        content = f.read()

    # Split by runs
    runs = re.split(r'RUN \d+:', content)[1:]  # Skip header

    for run_idx, run in enumerate(runs):
        # Extract description
        desc_match = re.search(r'Description:\s*(.+?)(?:\n|$)', run)
        if desc_match:
            description = desc_match.group(1).strip()
        else:
            description = ""
        descriptions.append(description)

        # Extract core lattice
        lattice_match = re.search(r'core_lattice:\s*(\[.*?\])\n', run, re.DOTALL)
        if lattice_match:
            lattice_str = lattice_match.group(1)
            # Clean up the string for ast.literal_eval
            lattice_str = re.sub(r'\s+', ' ', lattice_str)
            lattice_str = lattice_str.replace('\n', '')
            try:
                lattice_list = ast.literal_eval(lattice_str)
                lattice = np.array(lattice_list)
                lattices.append(lattice)
            except:
                print(f"Failed to parse lattice in RUN {run_idx + 1}: {lattice_str[:50]}...")
                lattices.append(np.array([]))
                continue
        else:
            lattices.append(np.array([]))
            continue

        # Extract k-effective
        k_eff_match = re.search(r'k-effective:\s*([\d.]+)', run)
        if k_eff_match:
            k_effectives.append(float(k_eff_match.group(1)))
        else:
            k_effectives.append(None)

        # Extract flux data AND energy groups
        flux_dict_raw = {}
        energy_dict = {}  # NEW: Store energy percentages

        # Pattern to match flux line with energy groups
        flux_pattern = r'(I_\d+) Flux ([\d.]+e[+-]\d+) \[([\d.]+)% thermal, ([\d.]+)% epithermal, ([\d.]+)% fast\]'
        flux_matches = re.findall(flux_pattern, run)

        for irr_pos, flux_val, thermal, epithermal, fast in flux_matches:
            flux_dict_raw[irr_pos] = float(flux_val)

            # Store energy percentages as fractions (0-1)
            thermal_frac = float(thermal) / 100.0
            epithermal_frac = float(epithermal) / 100.0
            fast_frac = float(fast) / 100.0

            # Validate and normalize if needed
            total = thermal_frac + epithermal_frac + fast_frac
            if abs(total - 1.0) > 0.005:  # More than 0.5% off
                all_warnings.append(f"RUN {run_idx + 1}: Energy fractions for {irr_pos} sum to {total:.3f}, normalizing to 1.0")
                thermal_frac /= total
                epithermal_frac /= total
                fast_frac /= total

            energy_dict[irr_pos] = {
                'thermal': thermal_frac,
                'epithermal': epithermal_frac,
                'fast': fast_frac
            }

        # Validate and normalize flux data
        if lattice.size > 0:  # Only validate if lattice was parsed successfully
            flux_dict_normalized, warnings = validate_irradiation_positions(lattice, flux_dict_raw)
            if warnings:
                all_warnings.extend([f"RUN {run_idx + 1}: {w}" for w in warnings])
            flux_data.append(flux_dict_normalized)
            energy_groups.append(energy_dict)
        else:
            flux_data.append(flux_dict_raw)
            energy_groups.append(energy_dict)

    # Print all warnings at the end
    if all_warnings:
        print("\nWarnings during parsing:")
        for warning in all_warnings:
            print(f"  - {warning}")
        print()

    return lattices, flux_data, k_effectives, descriptions, energy_groups

def create_reactor_data_excel(data_file_path: str, output_prefix: str = "parsed_data_", output_dir: str = None):
    """
    Create an Excel file with parsed reactor data.
    If 'test' is in output_prefix, only original configurations are included.
    Otherwise, all rotational symmetries are applied.

    Args:
        data_file_path: Path to the input data file
        output_prefix: Prefix for the output Excel filename (default: "parsed_data_")
        output_dir: Directory where the Excel file should be saved (default: current directory)

    Returns:
        str: Path to the created Excel file
    """
    # Parse the data
    lattices, flux_data, k_effectives, descriptions, energy_groups = parse_reactor_data(data_file_path)

    # Check if this is test data (no augmentation needed)
    is_test_data = 'test' in output_prefix.lower()

    # Prepare data for DataFrame
    excel_data = []

    for config_idx, (desc, lattice, k_eff, flux_dict, energy_dict) in enumerate(zip(descriptions, lattices, k_effectives, flux_data, energy_groups)):
        if lattice.size == 0:  # Skip if lattice parsing failed
            continue

        if is_test_data:
            # For test data, only include the original configuration
            row = {
                'Config_ID': config_idx,
                'Description': desc,
                'Lattice': str(lattice.tolist()),
                'k-eff': k_eff if k_eff is not None else 'N/A'
            }

            # Add flux values for I_1 through I_4
            flux_values = []
            for i in range(1, 5):
                irr_label = f'I_{i}'
                if irr_label in flux_dict:
                    row[irr_label] = flux_dict[irr_label]
                    flux_values.append(flux_dict[irr_label])
                else:
                    row[irr_label] = 'N/A'

            # Calculate average flux
            if flux_values:
                row['Average Flux'] = np.mean(flux_values)
            else:
                row['Average Flux'] = 'N/A'

            excel_data.append(row)
        else:
            # For training data, apply rotational symmetries to get all 8 variations
            augmented_configs = apply_rotational_symmetry(lattice, flux_dict, k_eff, energy_dict)

            # Handle both 3-tuple and 4-tuple returns from apply_rotational_symmetry
            for aug_idx, aug_data in enumerate(augmented_configs):
                if len(aug_data) == 4:
                    aug_lattice, aug_flux_dict, aug_k_eff, aug_energy_dict = aug_data
                else:
                    aug_lattice, aug_flux_dict, aug_k_eff = aug_data
                    aug_energy_dict = None
                # Determine the transformation type
                if aug_idx < 4:
                    transform = f"Rotation {aug_idx * 90}°"
                else:
                    transform = f"Rotation {(aug_idx - 4) * 90}° + Flip"

                row = {
                    'Original_Config_ID': config_idx,
                    'Description': desc,
                    'Transformation': transform,
                    'Lattice': str(aug_lattice.tolist()),
                    'k-eff': aug_k_eff if aug_k_eff is not None else 'N/A'
                }

                # For augmented data, we need to map the flux values back to I_1, I_2, etc.
                # based on the actual positions in the augmented lattice
                irr_label_map = {}
                for i in range(aug_lattice.shape[0]):
                    for j in range(aug_lattice.shape[1]):
                        if aug_lattice[i, j].startswith('I_'):
                            irr_label_map[aug_lattice[i, j]] = (i, j)

                # Add flux values for I_1 through I_4
                flux_values = []
                for i in range(1, 5):
                    irr_label = f'I_{i}'
                    if irr_label in irr_label_map:
                        # Find the flux value for this position
                        pos_i, pos_j = irr_label_map[irr_label]
                        pos_key = f"pos_{pos_i}_{pos_j}"
                        if pos_key in aug_flux_dict:
                            flux_val = aug_flux_dict[pos_key]
                        elif irr_label in flux_dict:  # Fallback to original
                            flux_val = flux_dict[irr_label]
                        else:
                            flux_val = 'N/A'
                        row[irr_label] = flux_val
                        if flux_val != 'N/A':
                            flux_values.append(flux_val)
                    else:
                        row[irr_label] = 'N/A'

                # Calculate average flux
                if flux_values:
                    row['Average Flux'] = np.mean(flux_values)
                else:
                    row['Average Flux'] = 'N/A'

                excel_data.append(row)

    # Create DataFrame
    df = pd.DataFrame(excel_data)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{output_prefix}{timestamp}.xlsx"

    # Determine full output path
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        excel_filepath = os.path.join(output_dir, excel_filename)
    else:
        excel_filepath = excel_filename

    # Create Excel writer
    with pd.ExcelWriter(excel_filepath, engine='openpyxl') as writer:
        if is_test_data:
            # For test data, just write the original configurations
            df.to_excel(writer, sheet_name='Test Configurations', index=False)
            ws = writer.sheets['Test Configurations']
            sheet_name = 'Test Configurations'
        else:
            # For training data, write all augmented data
            df.to_excel(writer, sheet_name='All Augmented Data', index=False)
            ws = writer.sheets['All Augmented Data']
            sheet_name = 'All Augmented Data'

        # Format headers
        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format scientific notation for flux columns
        from openpyxl.utils import get_column_letter

        flux_columns = ['I_1', 'I_2', 'I_3', 'I_4', 'Average Flux']
        for col_name in flux_columns:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(df) + 2):
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value != 'N/A' and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00E+00'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    # Limit lattice column width
                    if cell.column == df.columns.get_loc('Lattice') + 1:
                        max_length = min(50, max(max_length, len(str(cell.value))))
                    else:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # For training data, also create a summary sheet
        if not is_test_data:
            # Create summary sheet with just original configurations
            summary_data = []
            for config_id in df['Original_Config_ID'].unique():
                # Get the first row for this config (0° rotation, no flip)
                original_row = df[(df['Original_Config_ID'] == config_id) & (df['Transformation'] == 'Rotation 0°')].iloc[0]
                summary_data.append({
                    'Config_ID': config_id,
                    'Description': original_row['Description'],
                    'k-eff': original_row['k-eff'],
                    'I_1': original_row['I_1'],
                    'I_2': original_row['I_2'],
                    'I_3': original_row['I_3'],
                    'I_4': original_row['I_4'],
                    'Average Flux': original_row['Average Flux'],
                    'Augmented_Configs': 8  # Always 8 variations per original
                })

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Original Configurations', index=False)

            # Format summary sheet
            ws_summary = writer.sheets['Original Configurations']
            for cell in ws_summary[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Format flux columns in summary
            for col_name in flux_columns:
                if col_name in summary_df.columns:
                    col_idx = summary_df.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    for row in range(2, len(summary_df) + 2):
                        cell = ws_summary[f"{col_letter}{row}"]
                        if cell.value != 'N/A' and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00E+00'

            # Auto-adjust summary columns
            for column in ws_summary.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_summary.column_dimensions[column_letter].width = adjusted_width

    # Print summary based on data type
    if is_test_data:
        print(f"\nTest data saved to Excel file: {excel_filepath}")
        print(f"  - Sheet '{sheet_name}': Contains {len(df)} test configurations (no augmentation)")
    else:
        print(f"\nTraining data saved to Excel file: {excel_filepath}")
        print(f"  - Sheet 'All Augmented Data': Contains all {len(df)} configurations (with rotations/flips)")
        print(f"  - Sheet 'Original Configurations': Summary of {len(summary_df)} original configurations")

    return excel_filepath

def apply_rotational_symmetry(lattice, flux_dict, k_eff, energy_dict=None):
    """
    Apply 8-fold rotational symmetry augmentation to reactor configuration
    FIXED: Properly tracks flux values during rotation
    NOW: Also handles energy group data
    """
    augmented_data = []

    # Original configuration
    if energy_dict is not None:
        augmented_data.append((lattice.copy(), flux_dict.copy(), k_eff, energy_dict.copy()))
    else:
        augmented_data.append((lattice.copy(), flux_dict.copy(), k_eff))

    # Track positions of irradiation cells
    irr_positions = {}
    for i in range(lattice.shape[0]):
        for j in range(lattice.shape[1]):
            if lattice[i, j].startswith('I'):
                irr_positions[lattice[i, j]] = (i, j)

    # 90 degree rotation
    rotated_90 = np.rot90(lattice, k=1)
    flux_90 = rotate_flux_values(lattice, rotated_90, flux_dict, irr_positions, k=1)
    energy_90 = rotate_flux_values(lattice, rotated_90, energy_dict, irr_positions, k=1) if energy_dict else None
    if energy_dict is not None:
        augmented_data.append((rotated_90, flux_90, k_eff, energy_90))
    else:
        augmented_data.append((rotated_90, flux_90, k_eff))

    # 180 degree rotation
    rotated_180 = np.rot90(lattice, k=2)
    flux_180 = rotate_flux_values(lattice, rotated_180, flux_dict, irr_positions, k=2)
    energy_180 = rotate_flux_values(lattice, rotated_180, energy_dict, irr_positions, k=2) if energy_dict else None
    if energy_dict is not None:
        augmented_data.append((rotated_180, flux_180, k_eff, energy_180))
    else:
        augmented_data.append((rotated_180, flux_180, k_eff))

    # 270 degree rotation
    rotated_270 = np.rot90(lattice, k=3)
    flux_270 = rotate_flux_values(lattice, rotated_270, flux_dict, irr_positions, k=3)
    energy_270 = rotate_flux_values(lattice, rotated_270, energy_dict, irr_positions, k=3) if energy_dict else None
    if energy_dict is not None:
        augmented_data.append((rotated_270, flux_270, k_eff, energy_270))
    else:
        augmented_data.append((rotated_270, flux_270, k_eff))

    # Horizontal flip
    flipped_h = np.fliplr(lattice)
    flux_h = flip_flux_values(lattice, flipped_h, flux_dict, irr_positions, axis='horizontal')
    energy_h = flip_flux_values(lattice, flipped_h, energy_dict, irr_positions, axis='horizontal') if energy_dict else None
    if energy_dict is not None:
        augmented_data.append((flipped_h, flux_h, k_eff, energy_h))
    else:
        augmented_data.append((flipped_h, flux_h, k_eff))

    # Vertical flip
    flipped_v = np.flipud(lattice)
    flux_v = flip_flux_values(lattice, flipped_v, flux_dict, irr_positions, axis='vertical')
    energy_v = flip_flux_values(lattice, flipped_v, energy_dict, irr_positions, axis='vertical') if energy_dict else None
    if energy_dict is not None:
        augmented_data.append((flipped_v, flux_v, k_eff, energy_v))
    else:
        augmented_data.append((flipped_v, flux_v, k_eff))

    # Diagonal flip (transpose)
    transposed = lattice.T
    flux_t = transpose_flux_values(lattice, transposed, flux_dict, irr_positions)
    energy_t = transpose_flux_values(lattice, transposed, energy_dict, irr_positions) if energy_dict else None
    if energy_dict is not None:
        augmented_data.append((transposed, flux_t, k_eff, energy_t))
    else:
        augmented_data.append((transposed, flux_t, k_eff))

    # Anti-diagonal flip
    anti_diag = np.fliplr(lattice.T)
    flux_ad = anti_diagonal_flux_values(lattice, anti_diag, flux_dict, irr_positions)
    energy_ad = anti_diagonal_flux_values(lattice, anti_diag, energy_dict, irr_positions) if energy_dict else None
    if energy_dict is not None:
        augmented_data.append((anti_diag, flux_ad, k_eff, energy_ad))
    else:
        augmented_data.append((anti_diag, flux_ad, k_eff))

    return augmented_data

def rotate_flux_values(original_lattice, rotated_lattice, flux_dict, irr_positions, k):
    """
    Correctly map flux values after rotation
    k: number of 90-degree counter-clockwise rotations (1, 2, or 3)

    numpy.rot90 rotates counter-clockwise:
    k=1: 90° CCW  -> (i,j) becomes (n-1-j, i)
    k=2: 180° CCW -> (i,j) becomes (n-1-i, n-1-j)
    k=3: 270° CCW -> (i,j) becomes (j, n-1-i)
    """
    new_flux = {}
    n = original_lattice.shape[0]

    # For each irradiation position, find where it moved to
    for label, original_pos in irr_positions.items():
        i, j = original_pos

        # Calculate new position after k 90-degree counter-clockwise rotations
        if k == 1:  # 90 degrees CCW
            new_i, new_j = n - 1 - j, i
        elif k == 2:  # 180 degrees CCW
            new_i, new_j = n - 1 - i, n - 1 - j
        elif k == 3:  # 270 degrees CCW (= 90 CW)
            new_i, new_j = j, n - 1 - i
        else:
            new_i, new_j = i, j

        # Verify the label is at the expected position
        if rotated_lattice[new_i, new_j] == label:
            new_flux[label] = flux_dict[label]
        else:
            # This shouldn't happen if rotation is correct
            print(f"Warning: Label {label} not found at expected position after rotation")
            # Search for the label
            for ri in range(n):
                for rj in range(n):
                    if rotated_lattice[ri, rj] == label:
                        new_flux[label] = flux_dict[label]
                        break

    return new_flux


def flip_flux_values(original_lattice, flipped_lattice, flux_dict, irr_positions, axis):
    """
    Correctly map flux values after flipping
    axis: 'horizontal' or 'vertical'
    """
    new_flux = {}
    n = original_lattice.shape[0]

    for label, original_pos in irr_positions.items():
        i, j = original_pos

        if axis == 'horizontal':
            new_i, new_j = i, n - 1 - j
        else:  # vertical
            new_i, new_j = n - 1 - i, j

        # Verify and map
        if flipped_lattice[new_i, new_j] == label:
            new_flux[label] = flux_dict[label]
        else:
            # Search for the label
            for ri in range(n):
                for rj in range(n):
                    if flipped_lattice[ri, rj] == label:
                        new_flux[label] = flux_dict[label]
                        break

    return new_flux


def transpose_flux_values(original_lattice, transposed_lattice, flux_dict, irr_positions):
    """
    Correctly map flux values after transpose
    """
    new_flux = {}

    for label, original_pos in irr_positions.items():
        i, j = original_pos
        new_i, new_j = j, i  # Transpose swaps indices

        if transposed_lattice[new_i, new_j] == label:
            new_flux[label] = flux_dict[label]
        else:
            # Search for the label
            n = original_lattice.shape[0]
            for ri in range(n):
                for rj in range(n):
                    if transposed_lattice[ri, rj] == label:
                        new_flux[label] = flux_dict[label]
                        break

    return new_flux


def anti_diagonal_flux_values(original_lattice, anti_diag_lattice, flux_dict, irr_positions):
    """
    Correctly map flux values after anti-diagonal flip (transpose + horizontal flip)
    """
    new_flux = {}
    n = original_lattice.shape[0]

    for label, original_pos in irr_positions.items():
        i, j = original_pos
        # Anti-diagonal: first transpose then flip horizontally
        new_i, new_j = n - 1 - j, n - 1 - i

        if anti_diag_lattice[new_i, new_j] == label:
            new_flux[label] = flux_dict[label]
        else:
            # Search for the label
            for ri in range(n):
                for rj in range(n):
                    if anti_diag_lattice[ri, rj] == label:
                        new_flux[label] = flux_dict[label]
                        break

    return new_flux

def encode_lattice_for_prediction(lattice: np.ndarray) -> Tuple[np.ndarray, List[Tuple]]:
    """
    Encode a single lattice for prediction.
    Returns: (feature_vector, irradiation_positions)
    """
    feature_vec = []
    irr_positions = []

    # Cell type encoding
    for i in range(lattice.shape[0]):
        for j in range(lattice.shape[1]):
            cell = lattice[i, j]
            if cell == 'F':
                feature_vec.extend([1, 0, 0])
            elif cell == 'C':
                feature_vec.extend([0, 1, 0])
            elif cell.startswith('I'):
                feature_vec.extend([0, 0, 1])
                irr_positions.append((i, j))
            else:
                feature_vec.extend([0, 0, 0])

    # Position encoding
    for i in range(lattice.shape[0]):
        for j in range(lattice.shape[1]):
            feature_vec.extend([i/7, j/7])

    return np.array(feature_vec), irr_positions

# Add this utility function to both txt_to_data.py and model_tester.py

def get_sorted_flux_values(lattice, flux_dict, irr_positions):
    """
    Get flux values in consistent order (sorted by label: I_1, I_2, I_3, I_4)

    Args:
        lattice: The reactor lattice
        flux_dict: Dictionary of flux values (can have various key formats)
        irr_positions: List of (row, col) tuples for irradiation positions

    Returns:
        List of flux values in consistent order [I_1_flux, I_2_flux, I_3_flux, I_4_flux]
    """
    # Create mapping of labels to flux values
    label_to_flux = {}

    for row, col in irr_positions:
        label = lattice[row, col]  # e.g., 'I_1', 'I_2', etc.

        # Try different key formats that might be in flux_dict
        possible_keys = [
            label,                      # Direct label
            f"pos_{row}_{col}",        # Position-based key
            (row, col),                # Tuple key
        ]

        for key in possible_keys:
            if key in flux_dict:
                label_to_flux[label] = flux_dict[key]
                break

    # Extract values in sorted label order
    flux_values = []
    for i in range(1, 5):  # I_1 through I_4
        label = f'I_{i}'
        if label in label_to_flux:
            flux_values.append(label_to_flux[label])

    return flux_values


# UPDATED prepare_ml_data function for txt_to_data.py:

def prepare_ml_data(lattices: List[np.ndarray], flux_data: List[Dict], k_effectives: List[float], energy_groups: List[Dict] = None):
    """
    Prepare data for ML training with rotational augmentation.
    Returns: (X_features, y_flux_values, y_k_eff, irr_positions_list)
    """
    X_features = []
    y_flux_values = []
    y_k_eff = []
    irr_positions_list = []

    # Handle backward compatibility
    if energy_groups is None:
        energy_groups = [{}] * len(lattices)

    for lattice, flux_dict, k_eff, energy_dict in zip(lattices, flux_data, k_effectives, energy_groups):
        # Apply rotational symmetry
        augmented = apply_rotational_symmetry(lattice, flux_dict, k_eff, energy_dict)

        for aug_data in augmented:
            # Handle both 3 and 4 value returns
            if len(aug_data) == 4:
                aug_lattice, aug_flux, aug_k_eff, aug_energy = aug_data
            else:
                aug_lattice, aug_flux, aug_k_eff = aug_data
                aug_energy = {}

            # Convert lattice to feature vector
            feature_vec = []
            irr_positions = []

            for i in range(aug_lattice.shape[0]):
                for j in range(aug_lattice.shape[1]):
                    cell = aug_lattice[i, j]
                    # Encode cell type
                    if cell == 'F':
                        feature_vec.extend([1, 0, 0])
                    elif cell == 'C':
                        feature_vec.extend([0, 1, 0])
                    elif cell.startswith('I_'):
                        feature_vec.extend([0, 0, 1])
                        irr_positions.append((i, j))
                    else:
                        feature_vec.extend([0, 0, 0])

            # Add position encoding for each cell
            for i in range(aug_lattice.shape[0]):
                for j in range(aug_lattice.shape[1]):
                    # Normalized position features
                    feature_vec.extend([i/7, j/7])

            X_features.append(feature_vec)

            # Get flux values in consistent sorted order
            flux_values = get_sorted_flux_values(aug_lattice, aug_flux, irr_positions)

            y_flux_values.append(flux_values)
            y_k_eff.append(aug_k_eff)
            irr_positions_list.append(irr_positions)

    return (np.array(X_features), np.array(y_flux_values),
            np.array(y_k_eff), irr_positions_list)

def validate_irradiation_positions(lattice: np.ndarray, flux_dict: Dict) -> Tuple[Dict, List[str]]:
    """
    Validate and normalize irradiation positions.

    Returns:
        normalized_flux: Dict with consistent I_1 through I_n labels
        warnings: List of any issues found
    """
    warnings = []

    # Find all irradiation positions in the lattice
    irr_positions = []
    irr_labels = []
    for i in range(lattice.shape[0]):
        for j in range(lattice.shape[1]):
            cell = lattice[i, j]
            if cell.startswith('I_'):
                irr_positions.append((i, j))
                irr_labels.append(cell)

    # Sort labels to ensure consistent ordering
    irr_labels.sort()

    # Check if we have the expected I_1 through I_4
    expected_labels = [f'I_{i}' for i in range(1, 5)]

    if irr_labels != expected_labels:
        warnings.append(f"Non-standard irradiation labels: {irr_labels} (expected {expected_labels})")

    # Normalize the flux dictionary
    normalized_flux = {}
    missing_flux = []

    for label in irr_labels:
        if label in flux_dict:
            normalized_flux[label] = flux_dict[label]
        else:
            missing_flux.append(label)
            # Use a default value or interpolate
            normalized_flux[label] = 0.0  # Or np.nan to flag missing data

    if missing_flux:
        warnings.append(f"Missing flux values for: {missing_flux}")

    # If we have non-standard labels, remap them
    if len(irr_labels) != 4 or irr_labels != expected_labels:
        remapped_flux = {}
        for idx, label in enumerate(sorted(irr_labels)):
            new_label = f'I_{idx + 1}'
            if label in normalized_flux:
                remapped_flux[new_label] = normalized_flux[label]
        normalized_flux = remapped_flux
        warnings.append(f"Remapped labels: {irr_labels} -> {list(remapped_flux.keys())}")

    return normalized_flux, warnings
